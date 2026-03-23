"""Smart Log Analysis Engine — Clean, formatted, management-ready Excel output.

This module acts as an intelligent analysis layer between raw parsed data
and the final Excel report.  It:
  • Extracts every meaningful insight from each software's log data
  • Computes proper calculations (averages, trends, peaks, utilisation %)
  • Formats everything into colour-coded, easy-to-read Excel sheets
  • Produces a one-page Dashboard for quick executive overview

Output structure:
  1. Dashboard        — high-level status of every software at a glance
  2. Ansys            — peak usage, monthly trends, rolling averages
  3. CATIA            — denials, users affected, system health
  4. Cortona          — IN/OUT/DENIED per user, daily activity
  5. MATLAB           — service health, errors & warnings
  6. Creo             — license entitlement data
  7. License Template — editable entitlements table
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

# ======================================================================
# Constants & styling
# ======================================================================

_ILLEGAL_CHARS_RE = re.compile(
    r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f\ufffe\uffff]"
)

# Colour palette  (hex without #)
_WHITE        = "FFFFFF"
_GREY_LIGHT   = "F5F5F5"
_GREY_MED     = "E0E0E0"
_BLACK        = "000000"

# Per-software colour themes: (dark, medium, light)
_COLOURS = {
    "dashboard": ("1B5E20", "4CAF50", "E8F5E9"),
    "ansys":     ("1565C0", "42A5F5", "E3F2FD"),
    "catia":     ("4A148C", "AB47BC", "F3E5F5"),
    "cortona":   ("E65100", "FF9800", "FFF3E0"),
    "matlab":    ("004D40", "26A69A", "E0F2F1"),
    "creo":      ("B71C1C", "EF5350", "FFEBEE"),
    "template":  ("37474F", "78909C", "ECEFF1"),
}

_THIN = Border(
    left=Side("thin", "CCCCCC"), right=Side("thin", "CCCCCC"),
    top=Side("thin", "CCCCCC"), bottom=Side("thin", "CCCCCC"),
)


def _sanitize(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in out.select_dtypes(include=["object"]).columns:
        out[col] = out[col].apply(
            lambda v: _ILLEGAL_CHARS_RE.sub("", v) if isinstance(v, str) else v
        )
    return out


# ======================================================================
# Date-window average calculator
# ======================================================================

def _window_avgs(dates: pd.Series, values: pd.Series) -> dict:
    """Compute full-period + 5/6/12 month rolling averages.

    Important: we compute the rolling windows at *calendar-month* granularity
    to avoid bias when some days are missing.
    """
    base = {"Full Period": 0.0, "Last 5 Months": 0.0, "Last 6 Months": 0.0, "Last 12 Months": 0.0}
    if dates is None or values is None or len(values) == 0:
        return base

    dt_full = pd.to_datetime(dates, errors="coerce")
    mask_ok = dt_full.notna() & (dt_full.dt.year >= 1900)
    if not mask_ok.any():
        return base

    dt = dt_full[mask_ok]
    vals = pd.to_numeric(values[mask_ok], errors="coerce")
    vals = vals.fillna(0.0)

    # Full-period mean across all valid records
    base["Full Period"] = float(vals.mean()) if len(vals) else 0.0

    # Month-bucket means, then rolling across last N months
    month = dt.dt.to_period("M")
    by_month = pd.DataFrame({"month": month.astype(str), "value": vals}).groupby("month")[["value"]].mean().reset_index()
    if by_month.empty:
        return base

    # Ensure chronological ordering
    by_month["month_dt"] = pd.to_datetime(by_month["month"] + "-01", errors="coerce")
    by_month = by_month.sort_values("month_dt")

    def _last_n(n: int) -> float:
        sub = by_month.tail(n)
        return float(sub["value"].mean()) if not sub.empty else 0.0

    base["Last 5 Months"] = _last_n(5)
    base["Last 6 Months"] = _last_n(6)
    base["Last 12 Months"] = _last_n(12)
    return base


def _hour_table_0_23(hourly: pd.DataFrame, hour_col: str = "Hour") -> pd.DataFrame:
    """Ensure an hour-wise table has a full 0–23 set of rows.

    Any missing hours are filled with 0 for numeric columns.
    """
    if hourly is None or hourly.empty or hour_col not in hourly.columns:
        return pd.DataFrame({hour_col: list(range(24))})

    full = pd.DataFrame({hour_col: list(range(24))})
    out = full.merge(hourly, on=hour_col, how="left").fillna(0)
    for c in out.columns:
        if c == hour_col:
            continue
        out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0)
        if (out[c] % 1 == 0).all():
            out[c] = out[c].astype(int)
        else:
            out[c] = out[c].round(2)
    return out


def _hourly_averages_from_daily(hour_tbl: pd.DataFrame, date_col: str = "Date", hour_col: str = "Hour") -> pd.DataFrame:
    """Compute average-per-day for each hour bucket.

    Input shape expected: one row per (Date, Hour) with numeric metric columns.
    Output: one row per Hour with Avg <metric>/Day.
    """
    if hour_tbl is None or hour_tbl.empty or date_col not in hour_tbl.columns or hour_col not in hour_tbl.columns:
        return pd.DataFrame()

    metrics = [c for c in hour_tbl.columns if c not in {date_col, hour_col}]
    if not metrics:
        return pd.DataFrame()

    tmp = hour_tbl.copy()
    tmp = tmp[tmp[date_col].notna() & tmp[hour_col].notna()].copy()
    if tmp.empty:
        return pd.DataFrame()

    grp = tmp.groupby(hour_col, dropna=False)[metrics]
    avg = grp.mean(numeric_only=True).reset_index()
    avg.columns = [hour_col] + [f"Avg {c}/Day" for c in metrics]
    avg = _hour_table_0_23(avg, hour_col=hour_col)
    return avg


def _as_dt(series: pd.Series, *, inferred_year: Optional[int] = None) -> pd.Series:
    """Best-effort datetime parsing for mixed timestamp strings.

    If `inferred_year` is provided, also try to parse Cortona-style values like
    'MM/DD HH:MM' by prefixing the year.
    """
    ts = pd.to_datetime(series, format="mixed", errors="coerce")
    if inferred_year is None:
        return ts

    # Many Cortona logs are 'MM/DD HH:MM' without a year.
    need_fix = ts.isna() & series.notna()
    if not need_fix.any():
        return ts

    s = series.astype(str)
    s2 = s.where(~need_fix, other=(str(inferred_year) + "/" + s))
    ts2 = pd.to_datetime(s2, errors="coerce")
    return ts.fillna(ts2)


def _percentile(series: pd.Series, q: float) -> Optional[float]:
    """Safe percentile helper (q in [0,1])."""
    try:
        s = pd.to_numeric(series, errors="coerce").dropna()
        if s.empty:
            return None
        return float(s.quantile(q))
    except Exception:
        return None


def _build_sessions_from_out_in(
    events: pd.DataFrame,
    *,
    ts_col: str = "timestamp",
    action_col: str = "action",
    out_value: str = "OUT",
    in_value: str = "IN",
    group_cols: Optional[List[str]] = None,
) -> pd.DataFrame:
    """Reconstruct sessions from OUT/IN event streams (greedy pairing per group)."""
    if events is None or events.empty:
        return pd.DataFrame()
    if ts_col not in events.columns or action_col not in events.columns:
        return pd.DataFrame()

    if group_cols is None:
        group_cols = [c for c in ["user", "feature", "host"] if c in events.columns]

    work = events.copy()
    work["_ts"] = _as_dt(work[ts_col])
    work = work[work["_ts"].notna()].copy()
    if work.empty:
        return pd.DataFrame()

    keep_cols = [c for c in group_cols if c in work.columns]
    if not keep_cols:
        keep_cols = []

    sessions: List[dict] = []
    if keep_cols:
        iterator = work.sort_values("_ts").groupby(keep_cols, dropna=False)
    else:
        iterator = [((), work.sort_values("_ts"))]

    for keys, g in iterator:
        if keep_cols and not isinstance(keys, tuple):
            keys = (keys,)
        key_map = dict(zip(keep_cols, keys)) if keep_cols else {}

        stack: List[pd.Timestamp] = []
        for _, r in g.iterrows():
            act = r.get(action_col)
            ts = r.get("_ts")
            if act == out_value:
                stack.append(ts)
            elif act == in_value and stack:
                start = stack.pop(0)
                end = ts
                if pd.notna(start) and pd.notna(end) and end >= start:
                    dur_min = (end - start).total_seconds() / 60.0
                    sessions.append({**key_map, "start_ts": start, "end_ts": end, "duration_min": dur_min})

    if not sessions:
        return pd.DataFrame()

    out = pd.DataFrame(sessions)
    out["duration_min"] = pd.to_numeric(out["duration_min"], errors="coerce")
    return out


def _hourly_concurrency_from_sessions(
    sessions: pd.DataFrame,
    *,
    start_col: str = "start_ts",
    end_col: str = "end_ts",
) -> pd.DataFrame:
    """Compute hourly average + peak concurrency from session intervals (all sessions combined)."""
    if sessions is None or sessions.empty:
        return pd.DataFrame()
    if start_col not in sessions.columns or end_col not in sessions.columns:
        return pd.DataFrame()

    s = sessions.copy()
    s[start_col] = pd.to_datetime(s[start_col], errors="coerce")
    s[end_col] = pd.to_datetime(s[end_col], errors="coerce")
    s = s[s[start_col].notna() & s[end_col].notna() & (s[end_col] >= s[start_col])].copy()
    if s.empty:
        return pd.DataFrame()

    starts = s[[start_col]].rename(columns={start_col: "ts"})
    starts["delta"] = 1
    ends = s[[end_col]].rename(columns={end_col: "ts"})
    ends["delta"] = -1
    ev = pd.concat([starts, ends], ignore_index=True)
    ev["HourStart"] = ev["ts"].dt.floor("h")
    by_hour = ev.groupby("HourStart")["delta"].sum().sort_index().cumsum().reset_index(name="Concurrency")
    if by_hour.empty:
        return pd.DataFrame()

    by_hour["Date"] = by_hour["HourStart"].dt.strftime("%Y-%m-%d")
    by_hour["Hour"] = by_hour["HourStart"].dt.hour

    hh = by_hour.groupby("Hour", dropna=False).agg(
        Avg_Concurrent=("Concurrency", "mean"),
        Peak_Concurrent=("Concurrency", "max"),
        Days=("Date", "nunique"),
    ).reset_index().sort_values("Hour")
    hh = _hour_table_0_23(hh, hour_col="Hour")
    if "Avg_Concurrent" in hh.columns:
        hh["Avg_Concurrent"] = pd.to_numeric(hh["Avg_Concurrent"], errors="coerce").fillna(0).round(2)
    if "Peak_Concurrent" in hh.columns:
        hh["Peak_Concurrent"] = pd.to_numeric(hh["Peak_Concurrent"], errors="coerce").fillna(0).astype(int)
    if "Days" in hh.columns:
        hh["Days"] = pd.to_numeric(hh["Days"], errors="coerce").fillna(0).astype(int)
    return hh


def _data_quality_kpis(df: pd.DataFrame, key_cols: List[str], ts_col: str = "timestamp") -> pd.DataFrame:
    """Small, human-readable data-quality summary to show coverage and confidence."""
    if df is None or df.empty:
        return pd.DataFrame([
            {"Check": "Records", "Result": "0"},
            {"Check": "Timestamp coverage", "Result": "0%"},
        ])

    out = []
    out.append({"Check": "Records", "Result": f"{len(df):,}"})

    if ts_col in df.columns:
        ts = _as_dt(df[ts_col])
        cov = (ts.notna().mean() * 100.0) if len(ts) else 0.0
        out.append({"Check": "Timestamp coverage", "Result": f"{cov:.0f}%"})

    for c in key_cols:
        if c in df.columns:
            cov = (df[c].notna().mean() * 100.0) if len(df[c]) else 0.0
            out.append({"Check": f"{c.title()} populated", "Result": f"{cov:.0f}%"})

    return pd.DataFrame(out)


# ======================================================================
# Excel formatting helpers
# ======================================================================

def _put_title(ws, row: int, text: str, theme: str, ncols: int = 8) -> int:
    """Write a dark banner title row. Returns next row."""
    dark, _, _ = _COLOURS.get(theme, _COLOURS["dashboard"])
    fill = PatternFill("solid", dark, dark)
    font = Font("Calibri", 14, bold=True, color=_WHITE)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
    ws.cell(row=row, column=1, value=text)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 28
    return row + 1


def _put_subtitle(ws, row: int, text: str, ncols: int = 8) -> int:
    """Write an italic narrative line. Returns next row."""
    ws.cell(row=row, column=1, value=text)
    ws.cell(row=row, column=1).font = Font("Calibri", 10, italic=True, color="555555")
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 36
    return row + 1


def _put_section(ws, row: int, title: str, theme: str, ncols: int) -> int:
    """Write a coloured section header. Returns next row."""
    _, med, _ = _COLOURS.get(theme, _COLOURS["dashboard"])
    fill = PatternFill("solid", med, med)
    font = Font("Calibri", 11, bold=True, color=_WHITE)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
    ws.cell(row=row, column=1, value=title)
    if ncols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 22
    return row + 1


def _put_table(ws, row: int, df: pd.DataFrame, theme: str) -> int:
    """Write a DataFrame as a formatted table. Returns next row after a gap."""
    if df.empty:
        return row

    _, _, light = _COLOURS.get(theme, _COLOURS["dashboard"])
    df = _sanitize(df)
    ncols = len(df.columns)

    # Column headers
    hdr_fill = PatternFill("solid", light, light)
    hdr_font = Font("Calibri", 10, bold=True, color=_BLACK)
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=row, column=ci, value=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    row += 1

    # Data rows
    data_font = Font("Calibri", 10)
    alt_fill = PatternFill("solid", _GREY_LIGHT, _GREY_LIGHT)
    for ri, (_, data_row) in enumerate(df.iterrows()):
        for ci, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = data_font
            cell.border = _THIN
            if ri % 2 == 1:
                cell.fill = alt_fill
            if isinstance(val, float):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(val, int):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
        row += 1

    return row + 1  # gap


def _put_table_simple(ws, row: int, df: pd.DataFrame, theme: str) -> int:
    """Write a plain DataFrame table with a real header row.

    Intended for export sheets (pivots/external sharing) where users expect
    the first visible table to have columns like User/Hostname/Software.
    """
    if df is None or df.empty:
        ws.cell(row=row, column=1, value="(no rows)")
    # Body
    for i in range(len(df)):
        for j in range(1, len(df.columns) + 1):
            val = df.iloc[i, j - 1]
            cell = ws.cell(row=row + 1 + i, column=j, value=val)
        return row + 2

    header_font = Font("Calibri", 10, bold=True, color=_BLACK)
    body_font = Font("Calibri", 10)
    hdr_fill = PatternFill("solid", _GREY_LIGHT, _GREY_LIGHT)

    # Header
    for j, col in enumerate(df.columns, 1):
        cell = ws.cell(row=row, column=j, value=str(col))
        cell.font = header_font
        cell.fill = hdr_fill
        cell.border = _THIN
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body
    for i in range(len(df)):
        for j in range(1, len(df.columns) + 1):
            val = df.iloc[i, j - 1]
            cell = ws.cell(row=row + 1 + i, column=j, value=val)
            cell.font = body_font
            cell.border = _THIN
            if isinstance(val, float):
                cell.number_format = "0.00"
            elif isinstance(val, int):
                cell.number_format = "#,##0"

    return row + 1 + len(df) + 2


def _auto_width(ws, max_col: int | None = None) -> None:
    """Auto-fit all column widths."""
    for ci in range(1, (max_col or ws.max_column or 1) + 1):
        mx = 8
        letter = get_column_letter(ci)
        for ri in range(1, min((ws.max_row or 0) + 1, 500)):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                mx = max(mx, min(len(str(v)) + 3, 50))
        ws.column_dimensions[letter].width = mx


# ======================================================================
# Analysis engine — one function per software
# ======================================================================

def _analyse_ansys_peak(df: pd.DataFrame) -> List[dict]:
    """Deep analysis of Ansys Peak CSV data."""
    if df.empty:
        return []

    sections = []

    # --- Product Summary ---
    summ = df[df.get("record_type", pd.Series(dtype=str)) == "summary"].copy()
    if not summ.empty and "product" in summ.columns:
        tbl = summ[["product", "average_usage", "total_count"]].copy()
        tbl = tbl.sort_values("total_count", ascending=False).reset_index(drop=True)

        # Add utilisation insight
        if "average_usage" in tbl.columns:
            tbl["utilisation_pct"] = (tbl["average_usage"] * 100).round(1)

        tbl.columns = ["Product Name", "Average Usage", "Total Count", "Utilisation %"]
        sections.append({
            "title": "📊 License Peak Usage — Product Summary",
            "table": tbl,
        })

    # --- Monthly Breakdown ---
    monthly = df[df.get("record_type", pd.Series(dtype=str)) == "monthly"].copy()
    if not monthly.empty and "month_label" in monthly.columns:
        try:
            pivot = monthly.pivot_table(
                index="product", columns="month_label",
                values="monthly_average", aggfunc="first",
            ).reset_index()
            # Sort months chronologically
            month_cols = [c for c in pivot.columns if c != "product"]
            try:
                month_cols = sorted(month_cols, key=lambda m: pd.to_datetime(m, format="%b-%y"))
            except Exception:
                pass
            pivot = pivot[["product"] + month_cols]
            pivot = pivot.rename(columns={"product": "Product Name"})
            # Convert to % for readability
            for mc in month_cols:
                pivot[mc] = (pivot[mc] * 100).round(1)
            sections.append({
                "title": "📅 Monthly Peak Usage (% of license capacity)",
                "table": pivot,
            })
        except Exception:
            pass

    # --- Rolling Window Averages ---
    daily = df[df.get("record_type", pd.Series(dtype=str)) == "daily"].copy()
    if not daily.empty and "product" in daily.columns:
        rows = []
        for product, g in daily.groupby("product", dropna=False):
            # Use the reconstructed "date" column (YYYY-MM-DD)
            dates = g["date"] if "date" in g.columns else pd.Series(dtype=str)
            avgs = _window_avgs(dates, g["peak_usage"])
            row = {"Product Name": product}
            row.update({k: round(v * 100, 1) for k, v in avgs.items()})
            rows.append(row)
        if rows:
            win_df = pd.DataFrame(rows)
            win_df = win_df.sort_values("Full Period", ascending=False).reset_index(drop=True)
            sections.append({
                "title": "📈 Rolling Average Peak Usage (% of capacity)",
                "table": win_df,
            })

    # --- ★ Average Usage Over Time (computed from daily data) ---
    if not daily.empty and "product" in daily.columns and "date" in daily.columns:
        # Parse real dates for date-range calculations
        daily["date_dt"] = pd.to_datetime(daily["date"], errors="coerce")

        usage_rows = []
        for product, g in daily.groupby("product", dropna=False):
            vals = g["peak_usage"].dropna()
            valid_dates = g["date_dt"].dropna()
            n_days = len(vals)
            days_used = int((vals > 0).sum())
            days_idle = n_days - days_used
            avg_when_used = float(vals[vals > 0].mean()) if days_used > 0 else 0.0
            peak_val = float(vals.max()) if n_days > 0 else 0.0
            overall_avg = float(vals.mean()) if n_days > 0 else 0.0

            # Date range
            if not valid_dates.empty:
                d_min, d_max = valid_dates.min(), valid_dates.max()
                span_days = (d_max - d_min).days + 1
                date_range = f"{d_min.strftime('%d-%b-%Y')} to {d_max.strftime('%d-%b-%Y')}"
            else:
                span_days = n_days
                date_range = "N/A"

            usage_rows.append({
                "Product": product,
                "Date Range": date_range,
                "Calendar Days": span_days,
                "Days With Data": n_days,
                "Days In Use": days_used,
                "Days Idle": days_idle,
                "Utilisation Rate": f"{days_used / n_days * 100:.1f}%" if n_days else "0%",
                "Avg Peak (when used)": f"{avg_when_used * 100:.1f}%",
                "Max Peak": f"{peak_val * 100:.0f}%",
                "Overall Avg": f"{overall_avg * 100:.1f}%",
            })
        if usage_rows:
            sections.append({
                "title": "📊 Average Usage Over Time (computed from daily log data)",
                "table": pd.DataFrame(usage_rows),
            })

    return sections


def _analyse_ansys_lm(df: pd.DataFrame) -> List[dict]:
    """Analyse Ansys License Manager admin log."""
    if df.empty:
        return []

    work = df.copy()
    if "date" not in work.columns:
        work["date"] = work.get("timestamp", "").astype(str).str[:10]

    grp = work.groupby("date", dropna=False)
    overview = grp.agg(
        running=("action", lambda s: int((s == "LM_RUNNING").sum())),
        stopped=("action", lambda s: int((s == "LM_STOPPED").sum())),
        uploads=("action", lambda s: int((s == "UPLOAD_LICENSE_FILE").sum())),
        errors=("action", lambda s: int((s == "ERROR").sum())),
        total_events=("action", "size"),
    ).reset_index()
    overview.columns = ["Date", "LM Running", "LM Stopped", "License Uploads", "Errors", "Total Events"]

    sections = [{"title": "🖥️ License Manager — Daily Activity", "table": overview}]

    # Hour-wise activity (only if timestamps have a time component)
    if "timestamp" in work.columns:
        ts = pd.to_datetime(work["timestamp"], format="mixed", errors="coerce")
        hh = work.copy()
        hh["_ts"] = ts
        hh = hh[hh["_ts"].notna()].copy()
        if not hh.empty:
            hh["Hour"] = hh["_ts"].dt.hour
            hour_tbl = hh.groupby("Hour").agg(
                LM_Running=("action", lambda s: int((s == "LM_RUNNING").sum())),
                LM_Stopped=("action", lambda s: int((s == "LM_STOPPED").sum())),
                Errors=("action", lambda s: int((s == "ERROR").sum())),
                Uploads=("action", lambda s: int((s == "UPLOAD_LICENSE_FILE").sum())),
                Events=("action", "size"),
            ).reset_index().sort_values("Hour")
            hour_tbl = _hour_table_0_23(hour_tbl, hour_col="Hour")
            sections.append({"title": "Hourly Activity (0–23)", "table": hour_tbl})

            hh["Date"] = hh["_ts"].dt.strftime("%Y-%m-%d")
            per_day = hh.groupby(["Date", "Hour"]).agg(
                LM_Running=("action", lambda s: int((s == "LM_RUNNING").sum())),
                LM_Stopped=("action", lambda s: int((s == "LM_STOPPED").sum())),
                Errors=("action", lambda s: int((s == "ERROR").sum())),
                Uploads=("action", lambda s: int((s == "UPLOAD_LICENSE_FILE").sum())),
                Events=("action", "size"),
            ).reset_index()
            avg_hour = _hourly_averages_from_daily(per_day, date_col="Date", hour_col="Hour")
            if not avg_hour.empty:
                sections.append({"title": "Hourly Averages (per day)", "table": avg_hour})

    return sections


def _analyse_catia(df: pd.DataFrame) -> List[dict]:
    """Deep analysis of CATIA LicenseServer logs."""
    if df.empty:
        return []

    work = df.copy()
    sections = []

    if "timestamp" in work.columns:
        work["timestamp_dt"] = pd.to_datetime(
            work["timestamp"], errors="coerce", format="%Y/%m/%d %H:%M:%S:%f"
        )
    if "date" not in work.columns:
        work["date"] = work["timestamp"].str.slice(0, 10)
    else:
        # Defensive: mixed types (str/float/NaN) can crash min/max comparisons.
        work["date"] = work["date"].astype(str)
        work.loc[work["date"].isin({"nan", "NaT", "None"}), "date"] = pd.NA

    denials = work[work["action"] == "LICENSE_DENIED"].copy()

    # --- Top-level stats (like a KPI card) ---
    total_events = len(work)
    total_denials = len(denials)
    total_grants = int((work["action"] == "LICENSE_GRANT").sum()) if "action" in work.columns else 0
    total_usage = int(work["action"].isin(["LICENSE_GRANT", "LICENSE_TIMEOUT", "LICENSE_DETACHMENT"]).sum()) if "action" in work.columns else 0
    # (work['date'] is normalized above)
    if "date" in work.columns:
        _dates = (
            work["date"]
            .dropna()
            .astype(str)
            .map(lambda s: s.strip())
        )
        _dates = _dates[(_dates != "") & (~_dates.isin({"nan", "NaT", "None"}))]
        date_range = f"{_dates.min()} to {_dates.max()}" if not _dates.empty else "N/A"
    else:
        date_range = "N/A"
    servers = int(work["host"].nunique()) if "host" in work.columns else 0
    denial_users = int(denials["user"].nunique()) if not denials.empty and "user" in denials.columns else 0
    all_users = int(work[work["user"].notna()]["user"].nunique()) if "user" in work.columns else 0

    kpi = pd.DataFrame([
        {"Metric": "Log Date Range", "Value": date_range},
        {"Metric": "Total Events Parsed", "Value": f"{total_events:,}"},
        {"Metric": "Unique Users (all events)", "Value": str(all_users)},
        {"Metric": "Servers / Computers", "Value": str(servers)},
        {"Metric": "License Grants", "Value": f"{total_grants:,}"},
        {"Metric": "License Usage Events", "Value": f"{total_usage:,}"},
        {"Metric": "License Denials", "Value": f"{total_denials:,}"},
        {"Metric": "Users Affected by Denials", "Value": str(denial_users)},
    ])
    sections.append({"title": "📋 Key Statistics", "table": kpi})

    sections.append({
        "title": "Data Quality & Coverage",
        "table": _data_quality_kpis(work, ["user", "feature", "host"], ts_col="timestamp"),
    })

    # --- Denials by User (the table user wants most) ---
    if not denials.empty and "user" in denials.columns:
        user_grp = denials[denials["user"].notna()].groupby("user", dropna=False)
        by_user = user_grp.agg(
            total_denials=("action", "size"),
            features_denied=("feature", lambda s: ", ".join(sorted(s.dropna().unique()))),
            days_affected=("date", pd.Series.nunique),
            first_denial=("date", "min"),
            last_denial=("date", "max"),
        ).reset_index().sort_values("total_denials", ascending=False)
        by_user.columns = ["User Name", "Total Denials", "Features Denied",
                           "Days Affected", "First Denial", "Last Denial"]
        sections.append({"title": "👤 License Denials — By User", "table": by_user.reset_index(drop=True)})

    # --- Denials by Feature ---
    if not denials.empty and "feature" in denials.columns:
        feat_grp = denials[denials["feature"].notna()].groupby("feature", dropna=False)
        by_feat = feat_grp.agg(
            total_denials=("action", "size"),
            unique_users=("user", pd.Series.nunique),
            users_list=("user", lambda s: ", ".join(sorted({str(u) for u in s.dropna()}))),
            days_affected=("date", pd.Series.nunique),
        ).reset_index().sort_values("total_denials", ascending=False)
        by_feat.columns = ["Feature Name", "Total Denials", "Users Affected",
                           "Users List", "Days"]
        sections.append({"title": "🔑 License Denials — By Feature", "table": by_feat.reset_index(drop=True)})

    # --- Daily Denial Trend ---
    if not denials.empty:
        daily = denials.groupby("date", dropna=False).agg(
            denials=("action", "size"),
            users=("user", pd.Series.nunique),
            features=("feature", pd.Series.nunique),
        ).reset_index()
        daily.columns = ["Date", "Denials", "Unique Users", "Unique Features"]
        sections.append({"title": "📅 Daily Denial Trend", "table": daily})

        # --- Hour-of-day Denial Heatmap (counts by hour) ---
        if "timestamp" in denials.columns:
            # CATIA timestamps often look like "YYYY/MM/DD HH:MM:SS:ms".
            # Coerce robustly by normalizing the millisecond separator.
            _ts = denials["timestamp"].astype(str).str.replace(r"^(\d{4}/\d{2}/\d{2} \d{2}:\d{2}:\d{2}):(\d{1,4})$", r"\1.\2", regex=True)
            denials["ts_dt"] = pd.to_datetime(_ts, errors="coerce")
            by_hour = denials[denials["ts_dt"].notna()].copy()
            if not by_hour.empty:
                by_hour["Hour"] = by_hour["ts_dt"].dt.hour
                hour_tbl = by_hour.groupby("Hour").agg(
                    Denials=("action", "size"),
                    Unique_Users=("user", pd.Series.nunique),
                    Unique_Features=("feature", pd.Series.nunique),
                ).reset_index().sort_values("Hour")
                hour_tbl.columns = ["Hour (0-23)", "Denials", "Unique Users", "Unique Features"]
                sections.append({"title": "Hourly Denials (0–23)", "table": hour_tbl})

                # Hour-wise averages (per day)
                by_hour["Date"] = pd.to_datetime(by_hour["date"], errors="coerce").dt.strftime("%Y-%m-%d")
                per_day = by_hour.groupby(["Date", "Hour"]).agg(
                    Denials=("action", "size"),
                    Unique_Users=("user", pd.Series.nunique),
                    Unique_Features=("feature", pd.Series.nunique),
                ).reset_index()
                avg_hour = _hourly_averages_from_daily(per_day, date_col="Date", hour_col="Hour")
                if not avg_hour.empty:
                    sections.append({"title": "Hourly Averages (per day)", "table": avg_hour})

    # --- System Health ---
    sys_acts = ["SERVER_START", "SERVER_STOP", "SYSTEM_SUSPEND", "SYSTEM_RESUME", "UPLOAD_FAIL"]
    sys_events = work[work["action"].isin(sys_acts)]
    if not sys_events.empty:
        sys_grp = sys_events.groupby("date", dropna=False)
        sys_daily = sys_grp.agg(
            starts=("action", lambda s: int((s == "SERVER_START").sum())),
            stops=("action", lambda s: int((s == "SERVER_STOP").sum())),
            suspends=("action", lambda s: int((s == "SYSTEM_SUSPEND").sum())),
            resumes=("action", lambda s: int((s == "SYSTEM_RESUME").sum())),
            upload_fails=("action", lambda s: int((s == "UPLOAD_FAIL").sum())),
        ).reset_index()
        sys_daily.columns = ["Date", "Server Starts", "Server Stops",
                             "System Suspends", "System Resumes", "Upload Failures"]
        sections.append({"title": "⚙️ System Health — Daily", "table": sys_daily})

    # --- ★ Average Denials Over Time (rolling averages from daily data) ---
    if not denials.empty:
        denial_daily = denials.groupby("date", dropna=False).agg(
            count=("action", "size"),
            users=("user", pd.Series.nunique),
        ).reset_index()
        avg_rows = []
        for label, col in [("Daily Denials", "count"), ("Daily Users Affected", "users")]:
            avgs = _window_avgs(denial_daily["date"], denial_daily[col])
            row = {"Metric": label}
            row.update({k: round(v, 1) for k, v in avgs.items()})
            avg_rows.append(row)
        if avg_rows:
            sections.append({
                "title": "📊 Average Denials Over Time — Rolling Averages",
                "table": pd.DataFrame(avg_rows),
            })

        # Monthly denial summary
        denial_daily["date_dt"] = pd.to_datetime(denial_daily["date"], errors="coerce")
        denial_daily["month"] = denial_daily["date_dt"].dt.to_period("M").astype(str)
        monthly = denial_daily[denial_daily["month"].notna()].groupby("month").agg(
            total_denials=("count", "sum"),
            active_days=("count", "size"),
            avg_per_day=("count", "mean"),
            max_per_day=("count", "max"),
        ).reset_index()
        monthly["avg_per_day"] = monthly["avg_per_day"].round(1)
        monthly.columns = ["Month", "Total Denials", "Days with Denials",
                           "Avg Denials/Day", "Peak Denials/Day"]
        sections.append({
            "title": "📅 Monthly Denial Summary",
            "table": monthly,
        })

    # --- License Usage (Grant / TimeOut / Detachment) ---
    usage_actions = ["LICENSE_GRANT", "LICENSE_TIMEOUT", "LICENSE_DETACHMENT"]
    usage_events = work[work["action"].isin(usage_actions)]
    if not usage_events.empty:
        usage_users = int(usage_events["user"].nunique()) if "user" in usage_events.columns else 0
        usage_kpi = pd.DataFrame([
            {"Metric": "Total Usage Events", "Value": f"{len(usage_events):,}"},
            {"Metric": "Grants", "Value": f"{int((usage_events['action'] == 'LICENSE_GRANT').sum()):,}"},
            {"Metric": "TimeOuts", "Value": f"{int((usage_events['action'] == 'LICENSE_TIMEOUT').sum()):,}"},
            {"Metric": "Detachments", "Value": f"{int((usage_events['action'] == 'LICENSE_DETACHMENT').sum()):,}"},
            {"Metric": "Unique Users", "Value": str(usage_users)},
        ])
        sections.append({"title": "🟢 License Usage — Grants / Sessions", "table": usage_kpi})

        if "user" in usage_events.columns:
            by_u = usage_events[usage_events["user"].notna()].groupby("user").agg(
                grants=("action", lambda s: int((s == "LICENSE_GRANT").sum())),
                timeouts=("action", lambda s: int((s == "LICENSE_TIMEOUT").sum())),
                detachments=("action", lambda s: int((s == "LICENSE_DETACHMENT").sum())),
                active_days=("date", pd.Series.nunique),
            ).reset_index().sort_values("grants", ascending=False)
            by_u.columns = ["User", "Grants", "TimeOuts", "Detachments", "Active Days"]
            sections.append({"title": "🟢 License Usage — By User", "table": by_u.reset_index(drop=True)})

        if "session_minutes" in usage_events.columns:
            mins = pd.to_numeric(usage_events["session_minutes"], errors="coerce").dropna()
            if not mins.empty:
                p95 = _percentile(mins, 0.95)
                p99 = _percentile(mins, 0.99)
                dur = pd.DataFrame([
                    {"Metric": "Records with Session Minutes", "Value": f"{len(mins):,}"},
                    {"Metric": "Avg Session (min)", "Value": round(float(mins.mean()), 1)},
                    {"Metric": "P95 Session (min)", "Value": round(p95, 1) if p95 is not None else "N/A"},
                    {"Metric": "P99 Session (min)", "Value": round(p99, 1) if p99 is not None else "N/A"},
                    {"Metric": "Max Session (min)", "Value": round(float(mins.max()), 1)},
                ])
                sections.append({"title": "Session Duration Distribution", "table": dur})

    return sections


def _analyse_cortona(df: pd.DataFrame) -> List[dict]:
    """Deep analysis of Cortona RLM logs — richest user-level data."""
    if df.empty:
        return []

    work = df.copy()

    # Build a proper date column.  Cortona timestamps are "MM/DD HH:MM"
    # without a year.  Infer the year from the source_file path
    # (e.g. ".../03-01-2026/pgraphics.dlog").
    inferred_year = None
    if "source_file" in work.columns:
        import re as _re
        for sf in work["source_file"].dropna().unique():
            m = _re.search(r"(\d{4})", str(sf))
            if m:
                y = int(m.group(1))
                if 2020 <= y <= 2030:
                    inferred_year = y
                    break
    if inferred_year is None:
        inferred_year = datetime.now().year

    # Parse the date robustly: "11/23" → "2025-11-23"
    if "timestamp" in work.columns:
        work["date"] = work["timestamp"].astype(str).str.strip().str.split(" ").str[0]
    elif "date" not in work.columns:
        work["date"] = ""

    def _cortona_date(mmdd: str) -> Optional[str]:
        """Convert 'MM/DD' to 'YYYY-MM-DD' using inferred year."""
        try:
            parts = str(mmdd).strip().split("/")
            if len(parts) == 2:
                mm, dd = int(parts[0]), int(parts[1])
                return f"{inferred_year}-{mm:02d}-{dd:02d}"
        except (ValueError, TypeError):
            pass
        return None

    work["date_full"] = work["date"].apply(_cortona_date)
    sections = []

    user_events = work[work["user"].notna()].copy() if "user" in work.columns else pd.DataFrame()

    # --- KPI ---
    total_out = int((work["action"] == "OUT").sum()) if "action" in work.columns else 0
    total_in = int((work["action"] == "IN").sum()) if "action" in work.columns else 0
    total_denied = int((work["action"] == "DENIED").sum()) if "action" in work.columns else 0
    n_users = int(user_events["user"].nunique()) if not user_events.empty else 0
    n_features = int(work["feature"].nunique()) if "feature" in work.columns else 0
    n_days = int(work["date_full"].dropna().nunique())
    denial_rate = f"{total_denied / (total_out + total_denied) * 100:.1f}%" if (total_out + total_denied) > 0 else "0%"

    # Date range
    valid_dates = pd.to_datetime(work["date_full"], errors="coerce").dropna()
    if not valid_dates.empty:
        d_min = valid_dates.min().strftime("%d-%b-%Y")
        d_max = valid_dates.max().strftime("%d-%b-%Y")
        date_range_str = f"{d_min} to {d_max}"
    else:
        date_range_str = "N/A"

    kpi = pd.DataFrame([
        {"Metric": "Log Date Range", "Value": date_range_str},
        {"Metric": "Days of Activity", "Value": str(n_days)},
        {"Metric": "Total Checkouts (OUT)", "Value": f"{total_out:,}"},
        {"Metric": "Total Check-ins (IN)", "Value": f"{total_in:,}"},
        {"Metric": "Total Denials", "Value": f"{total_denied:,}"},
        {"Metric": "Denial Rate", "Value": denial_rate},
        {"Metric": "Active Users", "Value": str(n_users)},
        {"Metric": "Licensed Features", "Value": str(n_features)},
    ])
    sections.append({"title": "📋 Key Statistics", "table": kpi})

    sections.append({
        "title": "Data Quality & Coverage",
        "table": _data_quality_kpis(work, ["user", "feature", "host"], ts_col="timestamp"),
    })

    # --- User Usage Summary (matches screenshot: User, OUT, IN, DENIED, Total) ---
    if not user_events.empty:
        grp = user_events.groupby("user", dropna=False)
        usage = grp.agg(
            out=("action", lambda s: int((s == "OUT").sum())),
            in_=("action", lambda s: int((s == "IN").sum())),
            denied=("action", lambda s: int((s == "DENIED").sum())),
            features=("feature", lambda s: ", ".join(sorted(s.dropna().unique()))),
        ).reset_index()
        usage["total"] = usage["out"] + usage["in_"] + usage["denied"]
        usage = usage.sort_values("total", ascending=False)
        usage.columns = ["User Name", "OUT (Checkouts)", "IN (Check-ins)",
                         "DENIED", "Features Used", "Grand Total"]
        sections.append({"title": "👤 Server Logs Usage Summary — By User", "table": usage.reset_index(drop=True)})

    # --- Denials by Feature ---
    denied = work[work["action"] == "DENIED"] if "action" in work.columns else pd.DataFrame()
    if not denied.empty and "feature" in denied.columns:
        feat_grp = denied.groupby("feature", dropna=False)
        by_feat = feat_grp.agg(
            denials=("action", "size"),
            users=("user", lambda s: ", ".join(sorted({str(u) for u in s.dropna()}))),
            unique_users=("user", pd.Series.nunique),
        ).reset_index().sort_values("denials", ascending=False)
        by_feat.columns = ["Feature Name", "Total Denials", "Users Affected", "Unique Users"]
        sections.append({"title": "🔑 Denials — By Feature", "table": by_feat.reset_index(drop=True)})

    # --- Daily Activity ---
    daily_grp = work.groupby("date_full", dropna=False)
    daily = daily_grp.agg(
        out=("action", lambda s: int((s == "OUT").sum())),
        in_=("action", lambda s: int((s == "IN").sum())),
        denied=("action", lambda s: int((s == "DENIED").sum())),
        users=("user", pd.Series.nunique),
    ).reset_index()
    daily = daily[daily["date_full"].notna()]
    daily.columns = ["Date", "Checkouts", "Check-ins", "Denials", "Active Users"]
    sections.append({"title": "📅 Daily License Activity", "table": daily})

    # --- Hour-wise Activity (from timestamp time-of-day) ---
    # Cortona timestamps are like "MM/DD HH:MM". We can still extract the hour.
    if "timestamp" in work.columns:
        ts_str = work["timestamp"].astype(str)
        tod = ts_str.str.split(" ").str[-1]
        hour = pd.to_numeric(tod.str.split(":").str[0], errors="coerce")
        hh = work.copy()
        hh["Hour"] = hour
        hh = hh[hh["Hour"].notna()]
        if not hh.empty:
            hour_tbl = hh.groupby("Hour").agg(
                Checkouts=("action", lambda s: int((s == "OUT").sum())),
                Checkins=("action", lambda s: int((s == "IN").sum())),
                Denials=("action", lambda s: int((s == "DENIED").sum())),
                Active_Users=("user", pd.Series.nunique),
            ).reset_index().sort_values("Hour")
            hour_tbl["Hour"] = hour_tbl["Hour"].astype(int)
            hour_tbl.columns = ["Hour (0-23)", "OUT", "IN", "DENIED", "Active Users"]
            sections.append({"title": "Hourly Activity (0–23)", "table": hour_tbl})

            # Hour-wise averages (per day)
            hh["Date"] = hh["date_full"].astype(str)
            per_day = hh.groupby(["Date", "Hour"]).agg(
                OUT=("action", lambda s: int((s == "OUT").sum())),
                IN=("action", lambda s: int((s == "IN").sum())),
                DENIED=("action", lambda s: int((s == "DENIED").sum())),
                Users=("user", pd.Series.nunique),
            ).reset_index()
            avg_hour = _hourly_averages_from_daily(per_day, date_col="Date", hour_col="Hour")
            if not avg_hour.empty:
                sections.append({"title": "Hourly Averages (per day)", "table": avg_hour})

    # Session reconstruction + concurrency sizing (best-effort)
    if all(c in work.columns for c in ["timestamp", "action", "user", "feature"]):
        sess_src = work[work["action"].isin(["OUT", "IN"])].copy()
        # Ensure Cortona 'MM/DD HH:MM' timestamps become real datetimes.
        if not sess_src.empty:
            sess_src = sess_src.copy()
            sess_src["timestamp"] = _as_dt(sess_src["timestamp"], inferred_year=inferred_year)
        group_cols = ["user", "feature"] + (["host"] if "host" in sess_src.columns else [])
        sessions = _build_sessions_from_out_in(sess_src, group_cols=group_cols, ts_col="timestamp")
        if not sessions.empty:
            p95 = _percentile(sessions["duration_min"], 0.95)
            p99 = _percentile(sessions["duration_min"], 0.99)
            sizing = pd.DataFrame([
                {"Metric": "Reconstructed Sessions", "Value": f"{len(sessions):,}"},
                {"Metric": "Avg Session (min)", "Value": round(float(sessions["duration_min"].mean()), 1)},
                {"Metric": "P95 Session (min)", "Value": round(p95, 1) if p95 is not None else "N/A"},
                {"Metric": "P99 Session (min)", "Value": round(p99, 1) if p99 is not None else "N/A"},
                {"Metric": "Max Session (min)", "Value": round(float(sessions["duration_min"].max()), 1)},
            ])
            sections.append({"title": "Session Durations (reconstructed from OUT/IN)", "table": sizing})

            conc = _hourly_concurrency_from_sessions(sessions)
            if not conc.empty:
                conc = conc.rename(columns={
                    "Hour": "Hour (0-23)",
                    "Avg_Concurrent": "Avg Concurrent",
                    "Peak_Concurrent": "Peak Concurrent",
                })
                sections.append({
                    "title": "Note",
                    "table": pd.DataFrame([
                        {
                            "Item": "Concurrency is estimated",
                            "Details": "Computed from reconstructed OUT→IN sessions. If the log is missing IN events (or has resets), peak/avg concurrency may be understated.",
                        }
                    ]),
                })
                sections.append({"title": "Hourly Concurrency (estimated from sessions)", "table": conc})

    # --- Rolling Averages ---
    avg_rows = []
    raw_daily = work.groupby("date_full", dropna=False).agg(
        out=("action", lambda s: int((s == "OUT").sum())),
        in_=("action", lambda s: int((s == "IN").sum())),
        denied=("action", lambda s: int((s == "DENIED").sum())),
    ).reset_index()
    raw_daily = raw_daily[raw_daily["date_full"].notna()]
    for label, col in [("Daily Checkouts", "out"), ("Daily Check-ins", "in_"), ("Daily Denials", "denied")]:
        avgs = _window_avgs(raw_daily["date_full"], raw_daily[col])
        row = {"Metric": label}
        row.update(avgs)
        avg_rows.append(row)
    if avg_rows:
        sections.append({"title": "📈 Usage — Rolling Averages", "table": pd.DataFrame(avg_rows)})

    # --- ★ Average Usage Over Time — per user ---
    if not user_events.empty:
        grp2 = user_events.groupby("user", dropna=False)
        user_avg = grp2.agg(
            total_out=("action", lambda s: int((s == "OUT").sum())),
            total_in=("action", lambda s: int((s == "IN").sum())),
            total_denied=("action", lambda s: int((s == "DENIED").sum())),
            days_active=("date_full", lambda s: int(s.dropna().nunique())),
            features=("feature", lambda s: ", ".join(sorted(s.dropna().unique()))),
        ).reset_index()
        user_avg["total"] = user_avg["total_out"] + user_avg["total_in"] + user_avg["total_denied"]
        user_avg["avg_checkouts_per_day"] = (user_avg["total_out"] / user_avg["days_active"]).round(2)
        user_avg["denial_rate"] = user_avg.apply(
            lambda r: f"{r['total_denied'] / (r['total_out'] + r['total_denied']) * 100:.1f}%"
            if (r["total_out"] + r["total_denied"]) > 0 else "0%", axis=1
        )
        user_avg = user_avg.sort_values("total", ascending=False)
        user_avg.columns = [
            "User Name", "Total OUT", "Total IN", "Total DENIED",
            "Active Days", "Features Used", "Grand Total",
            "Avg Checkouts/Day", "Denial Rate",
        ]
        sections.append({
            "title": "📊 Average Usage Over Time — By User",
            "table": user_avg.reset_index(drop=True),
        })

    # --- ★ Average Usage Over Time — per feature ---
    if "feature" in work.columns:
        feat_events = work[work["feature"].notna()].copy()
        if not feat_events.empty:
            fgrp = feat_events.groupby("feature", dropna=False)
            feat_avg = fgrp.agg(
                out=("action", lambda s: int((s == "OUT").sum())),
                in_=("action", lambda s: int((s == "IN").sum())),
                denied=("action", lambda s: int((s == "DENIED").sum())),
                users=("user", pd.Series.nunique),
                days=("date_full", lambda s: int(s.dropna().nunique())),
            ).reset_index()
            feat_avg["avg_out_per_day"] = (feat_avg["out"] / feat_avg["days"]).round(2)
            feat_avg["denial_rate"] = feat_avg.apply(
                lambda r: f"{r['denied'] / (r['out'] + r['denied']) * 100:.1f}%"
                if (r["out"] + r["denied"]) > 0 else "0%", axis=1
            )
            feat_avg = feat_avg.sort_values("out", ascending=False)
            feat_avg.columns = [
                "Feature", "OUT", "IN", "DENIED",
                "Users", "Active Days", "Avg OUT/Day", "Denial Rate",
            ]
            sections.append({
                "title": "🔑 Average Usage Over Time — By Feature",
                "table": feat_avg.reset_index(drop=True),
            })

    return sections


def _analyse_cortona_admin(df: pd.DataFrame) -> List[dict]:
    """Analyse Cortona Admin server logs."""
    if df.empty:
        return []

    sections = []

    if "date" in df.columns:
        grp = df.groupby("date", dropna=False)
        overview = grp.agg(
            sessions=("action", lambda s: int((s == "ADMIN_START").sum())),
            restarts=("action", lambda s: int((s == "RLM_RESTART").sum())),
            act_fails=("action", lambda s: int((s == "ACTIVATION_FAILED").sum())),
            licenses=("action", lambda s: int((s == "ADD_LICENSE").sum())),
        ).reset_index()
        overview.columns = ["Date", "Admin Sessions", "RLM Restarts",
                            "Activation Failures", "Licenses Added"]
        sections.append({"title": "🛠️ Admin Server Activity", "table": overview})

    fails = df[df["action"] == "ACTIVATION_FAILED"] if "action" in df.columns else pd.DataFrame()
    if not fails.empty:
        cols = [c for c in ["timestamp", "host", "details"] if c in fails.columns]
        tbl = fails[cols].copy().rename(columns={"timestamp": "Timestamp", "host": "Server", "details": "Error"})
        sections.append({"title": "❌ Activation Failures — Detail", "table": tbl.head(20).reset_index(drop=True)})

    return sections


def _analyse_matlab(df: pd.DataFrame) -> List[dict]:
    """Analyse MATLAB ServiceHost logs."""
    if df.empty:
        return []

    sections = []
    work = df.copy()

    # KPI
    total = len(work)
    n_files = int(work["source_file"].nunique()) if "source_file" in work.columns else 0
    errors = int((work["level"] == "E").sum()) if "level" in work.columns else 0
    warnings = int((work["level"] == "W").sum()) if "level" in work.columns else 0
    n_days = int(work["date"].nunique()) if "date" in work.columns else 0
    health = "✅ Healthy" if errors == 0 else f"⚠️ {errors} errors found"

    kpi = pd.DataFrame([
        {"Metric": "Log Files Parsed", "Value": str(n_files)},
        {"Metric": "Days Covered", "Value": str(n_days)},
        {"Metric": "Total Events", "Value": f"{total:,}"},
        {"Metric": "Errors", "Value": str(errors)},
        {"Metric": "Warnings", "Value": str(warnings)},
        {"Metric": "Health Status", "Value": health},
    ])
    sections.append({"title": "📋 Key Statistics", "table": kpi})

    # Daily health
    if "date" in work.columns:
        grp = work.groupby("date", dropna=False)
        daily = grp.agg(
            events=("action", "size"),
            errors=("action", lambda s: int((s == "ERROR").sum())),
            warnings=("action", lambda s: int(s.astype(str).str.startswith("WARNING").sum())),
            components=("feature", pd.Series.nunique),
        ).reset_index()
        daily.columns = ["Date", "Total Events", "Errors", "Warnings", "Active Components"]
        sections.append({"title": "📅 Daily Service Health", "table": daily})

    # Hour-wise activity + hour-wise averages (per day)
    if "timestamp" in work.columns:
        hh = work.copy()
        hh["_ts"] = pd.to_datetime(hh["timestamp"], format="mixed", errors="coerce")
        hh = hh[hh["_ts"].notna()].copy()
        if not hh.empty:
            hh["Hour"] = hh["_ts"].dt.hour
            hour_tbl = hh.groupby("Hour").agg(
                Events=("action", "size"),
                Errors=("action", lambda s: int((s == "ERROR").sum())),
                Warnings=("action", lambda s: int(s.astype(str).str.startswith("WARNING").sum())),
                Components=("feature", pd.Series.nunique),
            ).reset_index().sort_values("Hour")
            hour_tbl = _hour_table_0_23(hour_tbl, hour_col="Hour")
            sections.append({"title": "Hourly Activity (0–23)", "table": hour_tbl})

            hh["Date"] = hh["_ts"].dt.strftime("%Y-%m-%d")
            per_day = hh.groupby(["Date", "Hour"]).agg(
                Events=("action", "size"),
                Errors=("action", lambda s: int((s == "ERROR").sum())),
                Warnings=("action", lambda s: int(s.astype(str).str.startswith("WARNING").sum())),
                Components=("feature", pd.Series.nunique),
            ).reset_index()
            avg_hour = _hourly_averages_from_daily(per_day, date_col="Date", hour_col="Hour")
            if not avg_hour.empty:
                sections.append({"title": "Hourly Averages (per day)", "table": avg_hour})

    # Errors & warnings detail
    if "level" in work.columns:
        issues = work[work["level"].isin(["E", "W"])].copy()
        if not issues.empty:
            tbl = issues[["timestamp", "level", "feature", "details"]].head(50).copy()
            tbl["level"] = tbl["level"].map({"E": "ERROR", "W": "WARNING"})
            tbl.columns = ["Timestamp", "Level", "Component", "Message"]
            sections.append({"title": "⚠️ Errors & Warnings — Detail", "table": tbl.reset_index(drop=True)})

    # Per-file summary
    if "source_file" in work.columns:
        fgrp = work.groupby("source_file")
        fsumm = fgrp.agg(
            events=("action", "size"),
            errors=("action", lambda s: int((s == "ERROR").sum())),
            warnings=("action", lambda s: int(s.astype(str).str.startswith("WARNING").sum())),
            first=("timestamp", "min"),
            last=("timestamp", "max"),
        ).reset_index()
        fsumm["source_file"] = fsumm["source_file"].apply(lambda p: Path(p).name)
        fsumm.columns = ["Log File", "Events", "Errors", "Warnings", "First Event", "Last Event"]
        sections.append({"title": "📁 Log Files Inventory", "table": fsumm})

    # ★ Average Usage Over Time — service activity rolling averages
    if "date" in work.columns:
        raw_daily = work.groupby("date", dropna=False).agg(
            events=("action", "size"),
            errors=("action", lambda s: int((s == "ERROR").sum())),
            warnings=("action", lambda s: int(s.astype(str).str.startswith("WARNING").sum())),
            components=("feature", pd.Series.nunique),
        ).reset_index()
        avg_rows = []
        for label, col in [("Daily Events", "events"), ("Daily Errors", "errors"),
                           ("Daily Warnings", "warnings"), ("Active Components", "components")]:
            avgs = _window_avgs(raw_daily["date"], raw_daily[col])
            row = {"Metric": label}
            row.update({k: round(v, 1) for k, v in avgs.items()})
            avg_rows.append(row)
        if avg_rows:
            sections.append({
                "title": "📊 Average Usage Over Time — Rolling Averages",
                "table": pd.DataFrame(avg_rows),
            })

    # ★ Component (feature) usage frequency
    if "feature" in work.columns:
        comp = work[work["feature"].notna()].groupby("feature").agg(
            events=("action", "size"),
            errors=("action", lambda s: int((s == "ERROR").sum())),
            warnings=("action", lambda s: int(s.astype(str).str.startswith("WARNING").sum())),
            days_active=("date", pd.Series.nunique),
        ).reset_index()
        comp["avg_events_per_day"] = (comp["events"] / comp["days_active"]).round(1)
        comp = comp.sort_values("events", ascending=False).head(20)
        comp.columns = ["Component", "Total Events", "Errors", "Warnings",
                        "Active Days", "Avg Events/Day"]
        sections.append({
            "title": "🔧 Average Usage Over Time — By Component",
            "table": comp.reset_index(drop=True),
        })

    return sections


def _analyse_catia_token(df: pd.DataFrame) -> List[dict]:
    """Analyse CATIA Token usage traces."""
    if df.empty:
        return []

    work = df.copy()
    work["file_name"] = work["source_file"].apply(lambda p: Path(p).name)
    if "timestamp" in work.columns:
        work["trace_date"] = work["timestamp"].str.slice(0, 10)

    sections = []

    # Coverage summary
    if "trace_date" in work.columns:
        cov = work.groupby("trace_date").agg(
            files=("file_name", "count"),
            servers=("host", pd.Series.nunique),
        ).reset_index()
        cov.columns = ["Date", "Token Files", "Active Servers"]
        sections.append({"title": "📁 Token File Coverage", "table": cov})

    return sections


def _analyse_catia_stats(df: pd.DataFrame) -> List[dict]:
    """Analyse CATIA usage stats inventory, usage events, and master data."""
    if df.empty:
        return []

    sections = []

    # ── 1. USAGE EVENTS (Grant / TimeOut / Detachment) ──
    usage = df[df.get("log_type", pd.Series(dtype=str)) == "license_usage_event"].copy()
    if not usage.empty:
        # Parse date
        usage["date_dt"] = pd.to_datetime(usage["date"], errors="coerce")

        # KPI for usage events
        n_grants = int((usage["action"] == "Grant").sum())
        n_timeouts = int((usage["action"] == "TimeOut").sum())
        n_detach = int((usage["action"] == "Detachment").sum())
        n_users = int(usage["user"].nunique())
        n_features = int(usage["feature"].nunique()) if "feature" in usage.columns else 0
        date_min = usage["date"].min()
        date_max = usage["date"].max()
        mins = usage["session_minutes"].dropna()
        avg_sess = f"{mins.mean():.0f} mins" if not mins.empty else "N/A"
        max_sess = f"{mins.max():.0f} mins" if not mins.empty else "N/A"

        kpi = pd.DataFrame([
            {"Metric": "Date Range", "Value": f"{date_min} to {date_max}"},
            {"Metric": "Total License Grants", "Value": f"{n_grants:,}"},
            {"Metric": "Total TimeOuts (auto-release)", "Value": f"{n_timeouts:,}"},
            {"Metric": "Total Detachments", "Value": f"{n_detach:,}"},
            {"Metric": "Active Users", "Value": str(n_users)},
            {"Metric": "Licensed Features", "Value": str(n_features)},
            {"Metric": "Avg Session Duration", "Value": avg_sess},
            {"Metric": "Max Session Duration", "Value": max_sess},
        ])
        sections.append({"title": "📋 CATIA Usage Events — Key Statistics", "table": kpi})

        # ★ Average Usage Over Time — per user
        user_grp = usage.groupby("user", dropna=False)
        user_usage = user_grp.agg(
            grants=("action", lambda s: int((s == "Grant").sum())),
            timeouts=("action", lambda s: int((s == "TimeOut").sum())),
            detachments=("action", lambda s: int((s == "Detachment").sum())),
            total_events=("action", "size"),
            days_active=("date", pd.Series.nunique),
            features=("feature", lambda s: ", ".join(sorted(s.dropna().unique()))),
            avg_session_min=("session_minutes", lambda s: round(s.dropna().mean(), 1) if s.dropna().any() else None),
            total_session_min=("session_minutes", lambda s: round(s.dropna().sum(), 1) if s.dropna().any() else None),
        ).reset_index()
        user_usage["avg_events_per_day"] = (user_usage["total_events"] / user_usage["days_active"]).round(1)
        user_usage = user_usage.sort_values("total_events", ascending=False)
        user_usage.columns = [
            "User (Machine ID)", "Grants", "TimeOuts", "Detachments",
            "Total Events", "Active Days", "Features Used",
            "Avg Session (min)", "Total Session (min)", "Avg Events/Day",
        ]
        sections.append({
            "title": "👤 Average Usage Over Time — By User",
            "table": user_usage.reset_index(drop=True),
        })

        # ★ Daily Usage — average sessions per day
        daily_grp = usage.groupby("date", dropna=False)
        daily = daily_grp.agg(
            grants=("action", lambda s: int((s == "Grant").sum())),
            timeouts=("action", lambda s: int((s == "TimeOut").sum())),
            detachments=("action", lambda s: int((s == "Detachment").sum())),
            active_users=("user", pd.Series.nunique),
            avg_session=("session_minutes", lambda s: round(s.dropna().mean(), 1) if s.dropna().any() else None),
        ).reset_index()
        daily.columns = ["Date", "Grants", "TimeOuts", "Detachments", "Active Users", "Avg Session (min)"]
        sections.append({"title": "📅 Daily Usage Activity", "table": daily})

        # ★ Rolling window average
        avg_rows = []
        for label, col in [("Daily Grants", "Grants"), ("Daily Users", "Active Users")]:
            avgs = _window_avgs(daily["Date"], daily[col])
            row = {"Metric": label}
            row.update({k: round(v, 1) for k, v in avgs.items()})
            avg_rows.append(row)
        if avg_rows:
            sections.append({
                "title": "📈 Average Usage Over Time — Rolling Averages",
                "table": pd.DataFrame(avg_rows),
            })

        # ★ Monthly usage summary
        usage_monthly = usage.copy()
        usage_monthly["month"] = usage_monthly["date_dt"].dt.to_period("M").astype(str)
        month_valid = usage_monthly[usage_monthly["month"].notna()].copy()
        month_grp = month_valid.groupby("month")
        monthly = month_grp.agg(
            grants=("action", lambda s: int((s == "Grant").sum())),
            timeouts=("action", lambda s: int((s == "TimeOut").sum())),
            users=("user", pd.Series.nunique),
            active_days=("date", pd.Series.nunique),
        ).reset_index()
        monthly["avg_daily_grants"] = (monthly["grants"] / monthly["active_days"].clip(lower=1)).round(1)
        monthly = monthly[["month", "grants", "timeouts", "users", "active_days", "avg_daily_grants"]]
        monthly.columns = ["Month", "Total Grants", "Total TimeOuts", "Active Users",
                           "Active Days", "Avg Daily Grants"]
        sections.append({"title": "📅 Monthly Usage Summary", "table": monthly})

        # ★ Feature usage breakdown
        if "feature" in usage.columns:
            feat_grp = usage[usage["feature"].notna()].groupby("feature")
            feat_usage = feat_grp.agg(
                grants=("action", lambda s: int((s == "Grant").sum())),
                timeouts=("action", lambda s: int((s == "TimeOut").sum())),
                detachments=("action", lambda s: int((s == "Detachment").sum())),
                users=("user", pd.Series.nunique),
                days=("date", pd.Series.nunique),
                avg_session=("session_minutes", lambda s: round(s.dropna().mean(), 1) if s.dropna().any() else None),
            ).reset_index()
            feat_usage["avg_grants_per_day"] = (feat_usage["grants"] / feat_usage["days"]).round(1)
            feat_usage = feat_usage.sort_values("grants", ascending=False)
            feat_usage.columns = [
                "Feature", "Grants", "TimeOuts", "Detachments",
                "Users", "Active Days", "Avg Session (min)", "Avg Grants/Day",
            ]
            sections.append({
                "title": "🔑 Average Usage Over Time — By Feature",
                "table": feat_usage.reset_index(drop=True),
            })

    # ── 2. MASTER / ENTITLEMENT DATA ──
    master = df[df.get("log_type", pd.Series(dtype=str)) == "license_master"].copy()
    if not master.empty:
        lic_cols = [c for c in ["lic_type", "lic_qty", "licence_start_date",
                                "licence_end_date", "licence_order_id",
                                "plant_no", "system", "serial_number"]
                    if c in master.columns]
        if lic_cols:
            tbl = master[lic_cols].copy()
            tbl.columns = [c.replace("_", " ").title() for c in tbl.columns]
            sections.append({
                "title": "📋 License Entitlements (from Master Data)",
                "table": tbl.reset_index(drop=True),
            })

    # ── 3. STAT FILE INVENTORY (original behaviour) ──
    stats = df[df.get("log_type", pd.Series(dtype=str)) == "license_usage_stat"].copy()
    if not stats.empty and "server_name" in stats.columns:
        srv = stats.groupby("server_name", dropna=False).agg(
            total=("file_name", "size"),
            daily=("file_type", lambda s: int((s == "daily_stat").sum())),
            monthly=("file_type", lambda s: int((s == "monthly_stat").sum())),
            earliest=("date", "min"),
            latest=("date", "max"),
        ).reset_index()
        srv = srv[srv["server_name"].notna()]
        srv.columns = ["Server", "Total Files", "Daily Stats", "Monthly Stats", "Earliest", "Latest"]
        sections.append({"title": "📁 Usage Stats Files — By Server", "table": srv})

    return sections


def _analyse_creo(df: pd.DataFrame) -> List[dict]:
    """Analyse Creo license data — extract entitlements, QTY, expiry tracking."""
    if df.empty:
        return []

    sections = []
    work = df.copy()

    # Identify the actual data rows (skip IMPORTANT/Retain headers, metadata rows)
    if "ptc_license_pack" in work.columns:
        content = work[work["ptc_license_pack"].notna()].copy()
        content = content[~content["ptc_license_pack"].astype(str).str.match(
            r"^(IMPORTANT|Retain|Customer|Host|PTC Host|$)", case=False
        )]
        # Detect the header row to find proper column mapping
        # The data has: Product ID | Product Description | Package ID | Package Description | QTY | LICENSE FILE END DATE | FEATURE NAME | FEATURE VERSION
        header_mask = content["ptc_license_pack"].astype(str).str.strip().eq("Product Description") \
                      | content.get("unnamed:_0", pd.Series(dtype=str)).astype(str).str.strip().eq("Product ID")
        if header_mask.any():
            hdr_idx = content[header_mask].index[0]
            content = content.loc[content.index > hdr_idx].copy()

        if not content.empty:
            # Map unnamed columns to proper names
            col_names = list(content.columns)
            proper = ["Product ID", "Product Description", "Package ID",
                       "Package Description", "QTY", "License End Date",
                       "Feature Name", "Feature Version"]
            mapping = {}
            for i, cn in enumerate(col_names[:len(proper)]):
                mapping[cn] = proper[i]
            content = content.rename(columns=mapping)

            # Keep only relevant columns
            keep = [c for c in proper if c in content.columns]
            if keep:
                tbl = content[keep].copy()
                # Parse QTY as numeric
                if "QTY" in tbl.columns:
                    tbl["QTY"] = pd.to_numeric(tbl["QTY"], errors="coerce")
                # Parse end dates
                if "License End Date" in tbl.columns:
                    tbl["License End Date"] = pd.to_datetime(
                        tbl["License End Date"], errors="coerce"
                    ).dt.strftime("%Y-%m-%d")

                tbl = tbl.dropna(subset=[keep[0]])  # drop rows where product id is NaN
                sections.append({"title": "📋 License Entitlements — Detail", "table": tbl.reset_index(drop=True)})

                # ★ License Summary — aggregated by Product
                if "Product Description" in tbl.columns and "QTY" in tbl.columns:
                    summ = tbl.groupby("Product Description", dropna=False).agg(
                        total_qty=("QTY", "sum"),
                        licenses=("Product ID", "count"),
                        features=("Feature Name", lambda s: ", ".join(sorted(s.dropna().unique())) if s.dropna().any() else ""),
                        earliest_expiry=("License End Date", "min"),
                        latest_expiry=("License End Date", "max"),
                    ).reset_index().sort_values("total_qty", ascending=False)
                    summ.columns = ["Product", "Total Licenses (QTY)", "Entries",
                                    "Features", "Earliest Expiry", "Latest Expiry"]
                    sections.append({
                        "title": "📊 License Summary — By Product",
                        "table": summ.reset_index(drop=True),
                    })

                # ★ Expiry tracking
                if "License End Date" in tbl.columns:
                    from datetime import datetime as _dt
                    today = _dt.now()
                    expiry = tbl[tbl["License End Date"].notna()].copy()
                    expiry["end_dt"] = pd.to_datetime(expiry["License End Date"], errors="coerce")
                    expiry["days_remaining"] = (expiry["end_dt"] - today).dt.days
                    expiry = expiry.sort_values("days_remaining")
                    exp_tbl = expiry[["Product Description", "Feature Name", "QTY",
                                      "License End Date", "days_remaining"]].copy()
                    exp_tbl.columns = ["Product", "Feature", "QTY", "Expiry Date", "Days Remaining"]
                    exp_tbl = exp_tbl[exp_tbl["Days Remaining"].notna()]
                    # Mark status
                    exp_tbl["Status"] = exp_tbl["Days Remaining"].apply(
                        lambda d: "❌ EXPIRED" if d < 0 else ("⚠️ <90 days" if d < 90 else
                                  ("🟡 <180 days" if d < 180 else "✅ OK"))
                    )
                    sections.append({
                        "title": "⏰ License Expiry Tracking",
                        "table": exp_tbl.reset_index(drop=True),
                    })

    # Customer/host metadata
    meta_rows = []
    if "ptc_license_pack" in work.columns:
        meta = work[work.get("unnamed:_0", pd.Series(dtype=str)).isin(
            ["Customer", "Customer Number", "Host Name", "PTC Host ID"]
        )]
        for _, r in meta.iterrows():
            meta_rows.append({"Field": r.get("unnamed:_0", ""), "Value": r.get("ptc_license_pack", "")})
    if meta_rows:
        sections.append({"title": "ℹ️ Customer & Host Information", "table": pd.DataFrame(meta_rows)})

    # File inventory
    if "source_file" in df.columns:
        info = []
        for sf in df["source_file"].unique():
            sub = df[df["source_file"] == sf]
            row = {"File": Path(sf).name, "Rows": len(sub), "Columns": len(sub.columns)}
            if "source_sheet" in sub.columns:
                row["Sheets"] = int(sub["source_sheet"].nunique())
            info.append(row)
        sections.append({"title": "📁 Source Files", "table": pd.DataFrame(info)})

    return sections


# ======================================================================
# Dashboard builder
# ======================================================================

def _build_dashboard(non_empty: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Executive dashboard — one row per software."""
    rows = []

    # Ansys
    if "ansys_peak" in non_empty:
        df = non_empty["ansys_peak"]
        summ = df[df.get("record_type", pd.Series(dtype=str)) == "summary"]
        n = int(summ["product"].nunique()) if "product" in summ.columns else 0
        top_avg = f"{summ['average_usage'].max():.0%}" if "average_usage" in summ.columns and not summ.empty else "N/A"
        rows.append({"Software": "Ansys", "Data Source": "Peak Usage CSV + LM Log",
                      "Key Finding": f"{n} products tracked, highest avg: {top_avg}",
                      "Status": "✅ Active", "Action Needed": "Review high-usage products"})
    elif "ansys" in non_empty:
        rows.append({"Software": "Ansys", "Data Source": "License Manager Log",
                      "Key Finding": f"{len(non_empty['ansys'])} admin events",
                      "Status": "ℹ️ Admin Only", "Action Needed": "Need FlexNet log for user data"})

    if "catia_license" in non_empty:
        df = non_empty["catia_license"]
        denials = df[df["action"] == "LICENSE_DENIED"] if "action" in df.columns else pd.DataFrame()
        n_den = len(denials)
        n_users = int(denials["user"].nunique()) if not denials.empty and "user" in denials.columns else 0
        status = "⚠️ Denials Found" if n_den else "✅ No Denials"
        action = f"{n_den} denials, {n_users} users affected — review capacity" if n_den else "None"
        rows.append({"Software": "CATIA", "Data Source": "LicenseServer Logs",
                      "Key Finding": f"{n_den:,} denials across {int(df['host'].nunique()) if 'host' in df.columns else 0} servers",
                      "Status": status, "Action Needed": action})

    if "cortona" in non_empty:
        df = non_empty["cortona"]
        outs = int((df["action"] == "OUT").sum()) if "action" in df.columns else 0
        dens = int((df["action"] == "DENIED").sum()) if "action" in df.columns else 0
        rate = f"{dens / (outs + dens) * 100:.1f}%" if (outs + dens) > 0 else "0%"
        status = "⚠️ Denials Found" if dens else "✅ Healthy"
        rows.append({"Software": "Cortona 3D", "Data Source": "RLM License Log",
                      "Key Finding": f"{outs} checkouts, {dens} denials ({rate} denial rate)",
                      "Status": status, "Action Needed": "Review denial patterns" if dens else "None"})

    if "nx" in non_empty:
        df = non_empty["nx"]
        outs = int((df.get("action", pd.Series(dtype=str)) == "OUT").sum())
        ins = int((df.get("action", pd.Series(dtype=str)) == "IN").sum())
        dens = int((df.get("action", pd.Series(dtype=str)) == "DENIED").sum())
        users = int(df["user"].nunique()) if "user" in df.columns else 0
        rate = f"{dens / (outs + dens) * 100:.1f}%" if (outs + dens) > 0 else "0%"
        status = "⚠️ Denials Found" if dens else "✅ Healthy"
        rows.append({"Software": "NX Siemens", "Data Source": "FlexLM/FlexNet debug log",
                      "Key Finding": f"{users} users, {outs} OUT, {ins} IN, {dens} DENIED ({rate})",
                      "Status": status, "Action Needed": "Review capacity & peak demand" if dens else "None"})

    if "matlab" in non_empty:
        df = non_empty["matlab"]
        errs = int((df["level"] == "E").sum()) if "level" in df.columns else 0
        warns = int((df["level"] == "W").sum()) if "level" in df.columns else 0
        status = "⚠️ Issues" if errs else "✅ Healthy"
        rows.append({"Software": "MATLAB", "Data Source": "ServiceHost Logs",
                      "Key Finding": f"{errs} errors, {warns} warnings",
                      "Status": status, "Action Needed": "Review errors" if errs else "None"})

    if "creo" in non_empty:
        rows.append({"Software": "Creo (PTC)", "Data Source": "License Excel",
                      "Key Finding": f"{len(non_empty['creo'])} rows imported",
                      "Status": "ℹ️ Static Data", "Action Needed": "Need usage logs for analysis"})

    return pd.DataFrame(rows) if rows else pd.DataFrame()

# ======================================================================
# Narrative builder  (plain English per software)
# ======================================================================

def _build_user_dashboard(non_empty: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Build a unified User Analytics Dashboard for ALL software types.

    Returns a per-user-per-day DataFrame with proper hour-based metrics.
    Only CATIA Usage has actual ``session_minutes``; for other software we
    estimate session hours from IN/OUT timestamp pairs where available,
    otherwise the hours column stays 0.
    """
    all_rows: list[pd.DataFrame] = []

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _parse_dates(series: pd.Series) -> pd.Series:
        return pd.to_datetime(series, format="mixed", dayfirst=False, errors="coerce")

    def _extract(df: pd.DataFrame, sw_name: str,
                 host_col: str | None = None,
                 hours_col: str | None = None,
                 action_col: str = "action",
                 feature_col: str = "feature") -> pd.DataFrame:
        """Generic per-event row extractor.

        *hours_col*  — column with duration in MINUTES (only ``catia_usage_stats``
                       has ``session_minutes``); converted to hours internally.
        """
        if df.empty or "user" not in df.columns or "date" not in df.columns:
            return pd.DataFrame()

        data = df[df["user"].notna() & (df["user"].astype(str).str.strip() != "")].copy()
        if data.empty:
            return pd.DataFrame()

        data["_dt"] = _parse_dates(data["date"])
        data = data[data["_dt"].notna()].copy()
        if data.empty:
            return pd.DataFrame()

        act = data[action_col].astype(str).str.upper() if action_col in data.columns else pd.Series("UNKNOWN", index=data.index)
        feat = data[feature_col].fillna("").astype(str) if feature_col in data.columns else pd.Series("", index=data.index)

        out = pd.DataFrame({
            "User": data["user"].astype(str).str.strip(),
            "Hostname": (data[host_col].astype(str).fillna("Unknown")
                         if host_col and host_col in data.columns else "Unknown"),
            "Software": sw_name,
            "Feature": feat.values,
            "Day": data["_dt"].dt.day_name(),
            "Date": data["_dt"].dt.strftime("%Y-%m-%d"),
            "Action": act.values,
            "Session Hrs": (pd.to_numeric(data[hours_col], errors="coerce").fillna(0) / 60.0).round(2)
                           if hours_col and hours_col in data.columns
                           else 0.0,
            "Is_Deny": act.str.contains(r"DENIED|DENY|FAIL", na=False).astype(int).values,
            "Is_In":   act.str.contains(r"^IN$|GRANT|CHECKOUT|DETACHMENT|TIMEOUT", na=False).astype(int).values,
            "Is_Out":  act.str.contains(r"^OUT$|CHECKIN", na=False).astype(int).values,
        })
        return out


    # ------------------------------------------------------------------
    # 1. CATIA Usage  (only source with real session_minutes)
    # ------------------------------------------------------------------
    if "catia_usage_stats" in non_empty:
        usage = non_empty["catia_usage_stats"]
        mask = usage.get("log_type", pd.Series(dtype=str)) == "license_usage_event"
        subset = usage[mask] if mask.any() else usage
        all_rows.append(_extract(subset, "CATIA Usage",
                                 host_col="server_name",
                                 hours_col="session_minutes"))

    # ------------------------------------------------------------------
    # 2. CATIA License  (Grant/TimeOut/Detachment + denials from logs)
    #    Compute session hours from Grant→TimeOut/Detachment timestamp pairs.
    # ------------------------------------------------------------------
    if "catia_license" in non_empty:
        cldf = non_empty["catia_license"].copy()

        # ── Compute session_minutes from Grant→TimeOut/Detachment pairs ──
        usage_actions = {"LICENSE_GRANT", "LICENSE_TIMEOUT", "LICENSE_DETACHMENT"}
        usage_mask = cldf["action"].isin(usage_actions) if "action" in cldf.columns else pd.Series(False, index=cldf.index)

        if usage_mask.any():
            usg = cldf[usage_mask].copy()
            # CATIA timestamps are "YYYY/MM/DD HH:MM:SS:mmm" — last : must be .
            usg["_ts_str"] = usg["timestamp"].astype(str).str.rsplit(":", n=1).str.join(".")
            usg["_ts"] = pd.to_datetime(usg["_ts_str"], format="mixed", errors="coerce")
            usg = usg[usg["_ts"].notna() & usg["user"].notna()].copy()
            usg = usg.sort_values(["user", "host", "_ts"])

            # Pair each Grant with the next TimeOut/Detachment for the same user+host
            mins_list: list[tuple] = []  # (index, minutes)
            open_grants: dict = {}  # (user, host) → (index, timestamp)
            for idx, row in usg.iterrows():
                key = (row["user"], row.get("host", ""))
                if row["action"] == "LICENSE_GRANT":
                    open_grants[key] = (idx, row["_ts"])
                elif row["action"] in ("LICENSE_TIMEOUT", "LICENSE_DETACHMENT"):
                    if key in open_grants:
                        grant_idx, grant_ts = open_grants.pop(key)
                        delta = (row["_ts"] - grant_ts).total_seconds() / 60.0
                        if 0 < delta < 1440:  # cap at 24 hours
                            mins_list.append((grant_idx, delta))
                            mins_list.append((idx, delta))

            cldf["session_minutes"] = 0.0
            for pair_idx, mins_val in mins_list:
                cldf.loc[pair_idx, "session_minutes"] = mins_val

        else:
            cldf["session_minutes"] = 0.0

        all_rows.append(_extract(cldf, "CATIA License",
                                 host_col="host",
                                 hours_col="session_minutes"))

    # ------------------------------------------------------------------
    # 3. Cortona 3D  (IN/OUT events; estimate hours from timestamp gaps)
    #    Cortona timestamps are "MM/DD HH:MM" without year.
    #    We infer the year from source_file mtime, else use previous year.
    # ------------------------------------------------------------------
    if "cortona" in non_empty:
        cdf = non_empty["cortona"].copy()

        # ── Infer year from file modification date ──
        from datetime import datetime as _dt_cls
        file_year = _dt_cls.now().year - 1  # default fallback
        file_month = _dt_cls.now().month
        if "source_file" in cdf.columns:
            src = cdf["source_file"].dropna().iloc[0] if cdf["source_file"].notna().any() else None
            if src:
                try:
                    mtime = Path(src).stat().st_mtime
                    fdt = _dt_cls.fromtimestamp(mtime)
                    file_year = fdt.year
                    file_month = fdt.month
                except Exception:
                    pass

        # ── Build proper date column: "MM/DD" → "YYYY/MM/DD" ──
        # If log month > file_month, the entry was likely from the prior year
        if "timestamp" in cdf.columns:
            raw_ts = cdf["timestamp"].astype(str).str.strip()
            mmdd = raw_ts.str.extract(r"^(\d{1,2})/(\d{1,2})", expand=True)
            mmdd.columns = ["mon", "day"]
            mmdd["mon"] = pd.to_numeric(mmdd["mon"], errors="coerce")
            mmdd["day"] = pd.to_numeric(mmdd["day"], errors="coerce")
            # Assign year: months > file_month → prior year (log wraps around year boundary)
            mmdd["year"] = file_year
            if file_month < 6:  # file from early in year → late months are from prior year
                mmdd.loc[mmdd["mon"] > file_month + 3, "year"] = file_year - 1
            cdf["date"] = mmdd.apply(
                lambda r: f"{int(r['year'])}/{int(r['mon']):02d}/{int(r['day']):02d}"
                if pd.notna(r["mon"]) and pd.notna(r["day"]) else "",
                axis=1,
            )

        # ── Estimate session hours from IN/OUT timestamp pairs ──
        cdf["session_minutes"] = 0.0
        if "timestamp" in cdf.columns and "user" in cdf.columns:
            # Parse timestamps with the proper year from the date column
            cdf["_ts"] = pd.to_datetime(cdf["date"] + " " +
                cdf["timestamp"].astype(str).str.strip().str.extract(r"(\d{1,2}:\d{2})", expand=False).fillna("00:00"),
                format="mixed", errors="coerce"
            )
            io = cdf[cdf["action"].isin(["IN", "OUT"]) & cdf["_ts"].notna()].sort_values(["user", "_ts"])
            sess_hrs: dict[tuple, float] = {}
            for usr, grp in io.groupby("user"):
                acts = grp[["_ts", "action"]].values
                last_in = None
                for ts, act in acts:
                    if act == "IN":
                        last_in = ts
                    elif act == "OUT" and last_in is not None:
                        delta = (pd.Timestamp(ts) - pd.Timestamp(last_in)).total_seconds() / 3600.0
                        if 0 < delta < 24:
                            day_key = pd.Timestamp(last_in).strftime("%Y/%m/%d")
                            sess_hrs[(usr, day_key)] = sess_hrs.get((usr, day_key), 0.0) + delta
                        last_in = None
            # Inject computed hours
            for (usr, day_str), hrs in sess_hrs.items():
                mask = (cdf["user"] == usr) & (cdf["date"] == day_str)
                idx = cdf.index[mask]
                if len(idx):
                    cdf.loc[idx[0], "session_minutes"] = hrs * 60
            cdf = cdf.drop(columns=["_ts"], errors="ignore")
        all_rows.append(_extract(cdf, "Cortona 3D", host_col="host",
                                 hours_col="session_minutes"))

    # ------------------------------------------------------------------
    # 4. Cortona Admin
    # ------------------------------------------------------------------
    if "cortona_admin" in non_empty:
        all_rows.append(_extract(non_empty["cortona_admin"], "Cortona Admin",
                                 host_col="host"))

    # ------------------------------------------------------------------
    # 5. Ansys Peak  (product-level daily peak usage, no user-level data)
    #    These are license-product rows, NOT real users. We prefix with
    #    "[Product]" so they are clearly distinguishable.
    # ------------------------------------------------------------------
    if "ansys_peak" in non_empty:
        pk = non_empty["ansys_peak"]
        daily = pk[pk.get("record_type", pd.Series(dtype=str)) == "daily"].copy()
        if not daily.empty and "product" in daily.columns and "date" in daily.columns:
            daily["user"] = "[Product] " + daily["product"].astype(str)
            daily["action"] = "PEAK_USAGE"
            daily["feature"] = daily["product"]
            daily["session_minutes"] = 0
            all_rows.append(_extract(daily, "Ansys Peak", hours_col=None))

    # ------------------------------------------------------------------
    # 6. MATLAB  (server-side log; estimate daily hours from timestamp span)
    # ------------------------------------------------------------------
    if "matlab" in non_empty:
        mdf = non_empty["matlab"].copy()
        if "timestamp" in mdf.columns and "user" in mdf.columns:
            mdf["_ts"] = pd.to_datetime(mdf["timestamp"], format="mixed", errors="coerce")
            mdf["_day"] = mdf["_ts"].dt.strftime("%Y-%m-%d")
            # For each user+day compute hours as (last_ts - first_ts)
            day_span = (mdf[mdf["user"].notna() & mdf["_ts"].notna()]
                        .groupby(["user", "_day"])["_ts"]
                        .agg(["min", "max"]))
            day_span["span_hrs"] = (day_span["max"] - day_span["min"]).dt.total_seconds() / 3600.0
            span_map = day_span["span_hrs"].to_dict()
            mdf["session_minutes"] = 0.0
            for (usr, day), hrs in span_map.items():
                if hrs > 0:
                    mask = (mdf["user"] == usr) & (mdf["_day"] == day)
                    idx = mdf.index[mask]
                    if len(idx):
                        mdf.loc[idx[0], "session_minutes"] = hrs * 60
            all_rows.append(_extract(mdf, "MATLAB", host_col="host",
                                     hours_col="session_minutes"))
        else:
            all_rows.append(_extract(mdf, "MATLAB", host_col="host"))

    # ------------------------------------------------------------------
    # 7. NX Siemens (FlexLM/FlexNet OUT/IN/DENIED)
    # ------------------------------------------------------------------
    if "nx" in non_empty:
        ndf = non_empty["nx"].copy()

        # Best-effort compute session_minutes from IN->OUT pairs, if we have a parseable timestamp
        ndf["session_minutes"] = 0.0
        if "timestamp" in ndf.columns and "user" in ndf.columns and "action" in ndf.columns:
            ndf["_ts"] = pd.to_datetime(ndf["timestamp"], format="mixed", errors="coerce")
            io = ndf[ndf["action"].isin(["IN", "OUT"]) & ndf["user"].notna() & ndf["_ts"].notna()].copy()
            if not io.empty:
                io = io.sort_values(["user", "host", "feature", "_ts"], kind="mergesort")
                open_in: dict[tuple, pd.Timestamp] = {}
                mins_map: dict[tuple, float] = {}
                for _, r in io.iterrows():
                    key = (str(r.get("user", "")).strip(), str(r.get("host", "")).strip(), str(r.get("feature", "")).strip())
                    act = str(r.get("action", "")).upper()
                    ts = r.get("_ts")
                    if pd.isna(ts):
                        continue
                    if act == "IN":
                        open_in[key] = ts
                    elif act == "OUT" and key in open_in:
                        start = open_in.pop(key)
                        delta_min = (pd.Timestamp(ts) - pd.Timestamp(start)).total_seconds() / 60.0
                        if 0 < delta_min < 1440:
                            day_key = pd.Timestamp(start).strftime("%Y-%m-%d")
                            mins_map[(key[0], day_key)] = mins_map.get((key[0], day_key), 0.0) + delta_min

                # Inject computed minutes into one row per user-day
                for (usr, day), mins in mins_map.items():
                    m = (ndf["user"].astype(str).str.strip() == usr) & (pd.to_datetime(ndf["date"], errors="coerce").dt.strftime("%Y-%m-%d") == day)
                    idx = ndf.index[m]
                    if len(idx):
                        ndf.loc[idx[0], "session_minutes"] = float(mins)

            ndf = ndf.drop(columns=["_ts"], errors="ignore")

        all_rows.append(_extract(ndf, "NX Siemens", host_col="host", hours_col="session_minutes"))

    # ------------------------------------------------------------------
    # Combine
    # ------------------------------------------------------------------
    if not all_rows:
        return pd.DataFrame()

    rows = [r for r in all_rows if r is not None and not r.empty]
    if not rows:
        # Nothing produced user-level rows (e.g. software with only admin/system logs).
        return pd.DataFrame()

    combined = pd.concat(rows, ignore_index=True)
    if combined.empty:
        return pd.DataFrame()

    # ------------------------------------------------------------------
    # Aggregate per User × Software × Date
    # ------------------------------------------------------------------
    grp = combined.groupby(["User", "Hostname", "Software", "Day", "Date"])
    dash = grp.agg(
        Checkouts=("Is_In", "sum"),
        Denials=("Is_Deny", "sum"),
        Events=("Action", "count"),
        Daily_Hrs=("Session Hrs", "sum"),
    ).reset_index()

    # ------------------------------------------------------------------
    # Rolling usage-hour averages (meaningful only where hours > 0)
    # ------------------------------------------------------------------
    dash["_dt"] = pd.to_datetime(dash["Date"], errors="coerce")
    # For year-less dates (e.g. Cortona "1-11-25"), fix to recent year
    year1_mask = dash["_dt"].notna() & (dash["_dt"].dt.year < 100)
    if year1_mask.any():
        from datetime import datetime
        current_year = datetime.now().year
        dash.loc[year1_mask, "_dt"] = dash.loc[year1_mask, "_dt"].apply(
            lambda d: d.replace(year=current_year if d.month <= datetime.now().month else current_year - 1)
        )
        dash.loc[year1_mask, "Date"] = dash.loc[year1_mask, "_dt"].dt.strftime("%Y-%m-%d")

    dash = dash[dash["_dt"].notna()].copy()
    dash = dash.sort_values(["Software", "User", "_dt"]).reset_index(drop=True)

    # Compute rolling averages safely per group (avoid reset_index misalignment)
    for col, days in [("Avg Hrs/Month", 30), ("Avg Hrs/3Mo", 90),
                      ("Avg Hrs/6Mo", 180), ("Avg Hrs/Year", 365)]:
        dash[col] = 0.0
        for (sw, usr), idx in dash.groupby(["Software", "User"]).groups.items():
            sub = dash.loc[idx].copy()
            rolled = (sub.set_index("_dt")["Daily_Hrs"]
                      .rolling(f"{days}D", min_periods=1).mean()
                      .round(2))
            dash.loc[idx, col] = rolled.values

    dash["Daily_Hrs"] = dash["Daily_Hrs"].round(2)
    dash = dash.drop(columns=["_dt"])

    dash = dash.rename(columns={
        "Daily_Hrs": "Daily Hrs",
    })

    cols = ["User", "Hostname", "Software", "Day", "Date",
            "Daily Hrs", "Avg Hrs/Month", "Avg Hrs/3Mo", "Avg Hrs/6Mo", "Avg Hrs/Year",
            "Checkouts", "Denials", "Events"]
    return dash[cols].sort_values(["Software", "User", "Date"], ascending=[True, True, False])


def _build_software_summary(dash_df: pd.DataFrame) -> pd.DataFrame:
    """Build a per-software consolidated summary from the user dashboard data."""
    if dash_df.empty:
        return pd.DataFrame()

    rows = []
    for sw in sorted(dash_df["Software"].unique()):
        s = dash_df[dash_df["Software"] == sw]
        unique_users = s["User"].nunique()
        active_days = s["Date"].nunique()
        total_events = int(s["Events"].sum())
        total_checkouts = int(s["Checkouts"].sum())
        total_denials = int(s["Denials"].sum())
        total_hrs = round(s["Daily Hrs"].sum(), 1)
        avg_daily = round(s["Daily Hrs"].mean(), 2) if total_hrs > 0 else 0.0
        peak_daily = round(s["Daily Hrs"].max(), 2) if total_hrs > 0 else 0.0
        denial_rate = round(total_denials / total_events * 100, 1) if total_events else 0.0
        dt_min = s["Date"].min()
        dt_max = s["Date"].max()

        rows.append({
            "Software": sw,
            "Unique Users": unique_users,
            "Active Days": active_days,
            "Total Events": total_events,
            "Checkouts": total_checkouts,
            "Denials": total_denials,
            "Denial %": denial_rate,
            "Total Hrs": total_hrs,
            "Avg Daily Hrs": avg_daily,
            "Peak Daily Hrs": peak_daily,
            "Date Range": f"{dt_min} → {dt_max}",
        })

    return pd.DataFrame(rows)


def _build_user_duration_summary(dash_df: pd.DataFrame) -> pd.DataFrame:
    """Per-user rollup across the time windows the UI wants.

    Output columns are designed for the main UI dashboard:
      - User, Hostname
      - Avg Hrs/Day (3/6/12 months)
      - Total Hrs (3/6/12 months)
      - Avg Hrs/Day (full period) + Total Hrs (full period)

    Notes:
      * Hours are based on the unified dashboard's "Daily Hrs".
      * "Avg Hrs/Day" is computed as (total_hours_in_window / active_days_in_window).
    """
    if dash_df is None or dash_df.empty:
        return pd.DataFrame()
    if "Date" not in dash_df.columns:
        return pd.DataFrame()

    tmp = dash_df.copy()
    tmp["_dt"] = pd.to_datetime(tmp["Date"], errors="coerce")
    tmp = tmp[tmp["_dt"].notna()].copy()
    if tmp.empty:
        return pd.DataFrame()

    max_dt = tmp["_dt"].max()

    def _win(days: int) -> pd.DataFrame:
        cutoff = max_dt - pd.Timedelta(days=days)
        w = tmp[tmp["_dt"] >= cutoff].copy()
        if w.empty:
            return pd.DataFrame(columns=["User", "Total Hrs", "Active Days", "Avg Hrs/Day"]).copy()
        g = (w.groupby("User")
             .agg(
                 Hostname=("Hostname", "first"),
                 Total_Hrs=("Daily Hrs", "sum"),
                 Active_Days=("Date", "nunique"),
             ).reset_index())
        g["Avg_Hrs_Day"] = (g["Total_Hrs"] / g["Active_Days"].clip(lower=1)).round(2)
        g["Total_Hrs"] = g["Total_Hrs"].round(1)
        g = g.rename(columns={"Total_Hrs": "Total Hrs", "Active_Days": "Active Days", "Avg_Hrs_Day": "Avg Hrs/Day"})
        return g[["User", "Hostname", "Total Hrs", "Active Days", "Avg Hrs/Day"]]

    full = _win(36500)  # effectively full period
    m3 = _win(90)
    m6 = _win(180)
    m12 = _win(365)

    out = full[["User", "Hostname"]].copy()
    out = out.merge(m3[["User", "Total Hrs", "Avg Hrs/Day"]], on="User", how="left", suffixes=("", ""))
    out = out.rename(columns={"Total Hrs": "Total Hrs (3 Mo)", "Avg Hrs/Day": "Avg Hrs/Day (3 Mo)"})

    out = out.merge(m6[["User", "Total Hrs", "Avg Hrs/Day"]], on="User", how="left")
    out = out.rename(columns={"Total Hrs": "Total Hrs (6 Mo)", "Avg Hrs/Day": "Avg Hrs/Day (6 Mo)"})

    out = out.merge(m12[["User", "Total Hrs", "Avg Hrs/Day"]], on="User", how="left")
    out = out.rename(columns={"Total Hrs": "Total Hrs (12 Mo)", "Avg Hrs/Day": "Avg Hrs/Day (12 Mo)"})

    full2 = full.copy()
    full2 = full2.rename(columns={"Total Hrs": "Total Hrs (Full)", "Avg Hrs/Day": "Avg Hrs/Day (Full)"})
    out = out.merge(full2[["User", "Total Hrs (Full)", "Avg Hrs/Day (Full)"]], on="User", how="left")

    for c in ["Total Hrs (3 Mo)", "Total Hrs (6 Mo)", "Total Hrs (12 Mo)", "Total Hrs (Full)"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).round(1)
    for c in ["Avg Hrs/Day (3 Mo)", "Avg Hrs/Day (6 Mo)", "Avg Hrs/Day (12 Mo)", "Avg Hrs/Day (Full)"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).round(2)

    return out.sort_values("Total Hrs (12 Mo)", ascending=False)


def _build_software_duration_summary(dash_df: pd.DataFrame) -> pd.DataFrame:
    """Per-software rollup across 3/6/12 months + full period for the UI."""
    if dash_df is None or dash_df.empty:
        return pd.DataFrame()
    if "Date" not in dash_df.columns or "Software" not in dash_df.columns:
        return pd.DataFrame()

    tmp = dash_df.copy()
    tmp["_dt"] = pd.to_datetime(tmp["Date"], errors="coerce")
    tmp = tmp[tmp["_dt"].notna()].copy()
    if tmp.empty:
        return pd.DataFrame()

    max_dt = tmp["_dt"].max()

    def _win(days: int) -> pd.DataFrame:
        cutoff = max_dt - pd.Timedelta(days=days)
        w = tmp[tmp["_dt"] >= cutoff].copy()
        if w.empty:
            return pd.DataFrame(columns=["Software", "Total Hrs", "Active Days", "Unique Users", "Avg Hrs/Day"]).copy()
        g = (w.groupby("Software")
             .agg(
                 Total_Hrs=("Daily Hrs", "sum"),
                 Active_Days=("Date", "nunique"),
                 Unique_Users=("User", "nunique"),
             ).reset_index())
        g["Avg_Hrs_Day"] = (g["Total_Hrs"] / g["Active_Days"].clip(lower=1)).round(2)
        g["Total_Hrs"] = g["Total_Hrs"].round(1)
        g = g.rename(columns={"Total_Hrs": "Total Hrs", "Active_Days": "Active Days", "Unique_Users": "Unique Users", "Avg_Hrs_Day": "Avg Hrs/Day"})
        return g[["Software", "Unique Users", "Total Hrs", "Active Days", "Avg Hrs/Day"]]

    full = _win(36500)
    m3 = _win(90)
    m6 = _win(180)
    m12 = _win(365)

    out = full[["Software", "Unique Users"]].copy()
    out = out.merge(m3[["Software", "Total Hrs", "Avg Hrs/Day"]], on="Software", how="left")
    out = out.rename(columns={"Total Hrs": "Total Hrs (3 Mo)", "Avg Hrs/Day": "Avg Hrs/Day (3 Mo)"})
    out = out.merge(m6[["Software", "Total Hrs", "Avg Hrs/Day"]], on="Software", how="left")
    out = out.rename(columns={"Total Hrs": "Total Hrs (6 Mo)", "Avg Hrs/Day": "Avg Hrs/Day (6 Mo)"})
    out = out.merge(m12[["Software", "Total Hrs", "Avg Hrs/Day"]], on="Software", how="left")
    out = out.rename(columns={"Total Hrs": "Total Hrs (12 Mo)", "Avg Hrs/Day": "Avg Hrs/Day (12 Mo)"})

    full2 = full.rename(columns={"Total Hrs": "Total Hrs (Full)", "Avg Hrs/Day": "Avg Hrs/Day (Full)"})
    out = out.merge(full2[["Software", "Total Hrs (Full)", "Avg Hrs/Day (Full)"]], on="Software", how="left")

    for c in ["Total Hrs (3 Mo)", "Total Hrs (6 Mo)", "Total Hrs (12 Mo)", "Total Hrs (Full)"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).round(1)
    for c in ["Avg Hrs/Day (3 Mo)", "Avg Hrs/Day (6 Mo)", "Avg Hrs/Day (12 Mo)", "Avg Hrs/Day (Full)"]:
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0).round(2)

    return out.sort_values("Total Hrs (12 Mo)", ascending=False)


# ======================================================================
# Cross-software Utilisation (for UI)
# ======================================================================

def _safe_mean(series: pd.Series) -> float:
    s = pd.to_numeric(series, errors="coerce").dropna()
    return float(s.mean()) if not s.empty else 0.0


def _safe_max(series: pd.Series) -> float:
    s = pd.to_numeric(series, errors="coerce").dropna()
    return float(s.max()) if not s.empty else 0.0


def _build_utilisation_summary(non_empty: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Build a consistent Utilisation table across all software.

    Output columns:
      Software, Metric, Value, Basis

    Notes:
      - Where OUT/IN session data exists, utilisation is shown as Avg Daily Hours.
      - Where only peak/capacity ratios exist (Ansys Peak), utilisation is shown as Avg Peak %.
      - Where neither exists, we fall back to activity (events/day) with clear labeling.
    """
    rows: list[dict] = []
    if not non_empty:
        return pd.DataFrame(columns=["Software", "Metric", "Value", "Basis"])

    # 1) If we can build the user dashboard, prefer hour-based utilisation.
    try:
        dash = _build_user_dashboard(non_empty)
    except Exception:
        dash = pd.DataFrame()

    if dash is not None and not dash.empty and "Daily Hrs" in dash.columns:
        sw = (dash.groupby("Software").agg(
            Total_Hrs=("Daily Hrs", "sum"),
            Active_Days=("Date", "nunique"),
            Avg_Daily_Hrs=("Daily Hrs", "mean"),
            Peak_Daily_Hrs=("Daily Hrs", "max"),
            Users=("User", "nunique"),
        ).reset_index())
        sw["Utilisation"] = (sw["Total_Hrs"] / sw["Active_Days"].clip(lower=1)).round(2)
        for _, r in sw.iterrows():
            rows.append({
                "Software": str(r["Software"]),
                "Metric": "Avg Hrs/Day",
                "Value": float(r["Utilisation"]),
                "Basis": "OUT/IN sessions / inferred session-hours (where available)",
            })

    # 2) Ansys Peak: peak_usage & average_usage are already ratios of capacity.
    if "ansys_peak" in non_empty:
        ap = non_empty["ansys_peak"].copy()
    # Prefer daily rows if they have ratios; otherwise fall back to the summary rows.
        daily = ap[ap.get("record_type", pd.Series(dtype=str)) == "daily"].copy()
        # Some parser versions expose these as *_ratio; support both.
        avg_col = "average_usage" if "average_usage" in daily.columns else (
            "average_usage_ratio" if "average_usage_ratio" in daily.columns else None
        )
        peak_col = "peak_usage" if "peak_usage" in daily.columns else (
            "peak_usage_ratio" if "peak_usage_ratio" in daily.columns else None
        )
        if not daily.empty and avg_col and peak_col:
            avg_pct = _safe_mean(daily[avg_col]) * 100
            peak_pct = _safe_max(daily[peak_col]) * 100
            rows.append({
                "Software": "Ansys (Peak)",
                "Metric": "Avg Usage %",
                "Value": round(avg_pct, 1),
                "Basis": f"FlexNet Peak CSV {avg_col} (ratio of capacity)",
            })
            rows.append({
                "Software": "Ansys (Peak)",
                "Metric": "Max Peak %",
                "Value": round(peak_pct, 0),
                "Basis": f"FlexNet Peak CSV {peak_col} (ratio of capacity)",
            })

        # Summary fallback (works even when date reconstruction isn't possible)
        summ = ap[ap.get("record_type", pd.Series(dtype=str)) == "summary"].copy()
        if not summ.empty and "average_usage" in summ.columns:
            avg_pct2 = _safe_mean(summ["average_usage"]) * 100
            # Avoid duplicating if we already added Avg Usage % from daily.
            if not any(r.get("Software") == "Ansys (Peak)" and r.get("Metric") == "Avg Usage %" for r in rows):
                rows.append({
                    "Software": "Ansys (Peak)",
                    "Metric": "Avg Usage %",
                    "Value": round(avg_pct2, 1),
                    "Basis": "FlexNet Peak CSV summary Average (ratio of capacity)",
                })

    # 3) Fallbacks: show activity intensity if hours/% not available.
    for key, df in non_empty.items():
        sw_name = {
            "ansys": "Ansys (LM)",
            "ansys_peak": "Ansys (Peak)",
            "catia_license": "CATIA (LicenseServer)",
            "catia_token": "CATIA (Token)",
            "catia_usage_stats": "CATIA (Usage Stats)",
            "cortona": "Cortona",
            "cortona_admin": "Cortona (Admin)",
            "creo": "Creo",
            "matlab": "MATLAB",
            "nx": "NX",
        }.get(key, key)

        # If already present as hour-based utilisation, don't add a confusing fallback.
        if any(r.get("Software") == sw_name and r.get("Metric") in ("Avg Hrs/Day", "Avg Usage %") for r in rows):
            continue

        if df is None or df.empty:
            continue

        # Try to compute events/day.
        date_col = None
        for c in ("date", "Date", "date_full"):
            if c in df.columns:
                date_col = c
                break
        if date_col is None:
            continue

        days = int(pd.to_datetime(df[date_col], errors="coerce").dropna().dt.date.nunique())
        events = int(len(df))
        if days <= 0:
            continue
        rows.append({
            "Software": sw_name,
            "Metric": "Events/Day",
            "Value": round(events / days, 1),
            "Basis": "Fallback (no session-hours/capacity ratio available)",
        })

    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(columns=["Software", "Metric", "Value", "Basis"])
    return out.sort_values(["Software", "Metric"]).reset_index(drop=True)


def _build_user_hour_summary(dash_df: pd.DataFrame) -> pd.DataFrame:
    """Per-user hour summary across all software — the consolidated view.

    Shows each user's total hours, avg daily hours, active days, denials
    and the software they used.  Only includes users that have non-zero
    hour data OR significant event counts.
    """
    if dash_df.empty:
        return pd.DataFrame()

    user_sw = (dash_df.groupby("User")
               .agg(
                   Software=("Software", lambda x: ", ".join(sorted(x.unique()))),
                   Hostname=("Hostname", "first"),
                   Active_Days=("Date", "nunique"),
                   Total_Events=("Events", "sum"),
                   Total_Checkouts=("Checkouts", "sum"),
                   Total_Denials=("Denials", "sum"),
                   Total_Hrs=("Daily Hrs", "sum"),
                   Avg_Daily_Hrs=("Daily Hrs", "mean"),
                   Peak_Daily_Hrs=("Daily Hrs", "max"),
                   First_Seen=("Date", "min"),
                   Last_Seen=("Date", "max"),
               ).reset_index())

    user_sw["Total_Hrs"] = user_sw["Total_Hrs"].round(1)
    user_sw["Avg_Daily_Hrs"] = user_sw["Avg_Daily_Hrs"].round(2)
    user_sw["Peak_Daily_Hrs"] = user_sw["Peak_Daily_Hrs"].round(2)
    user_sw["Denial_Rate"] = (user_sw["Total_Denials"] / user_sw["Total_Events"] * 100).round(1).fillna(0)

    user_sw = user_sw.rename(columns={
        "Active_Days": "Active Days",
        "Total_Events": "Events",
        "Total_Checkouts": "Checkouts",
        "Total_Denials": "Denials",
        "Total_Hrs": "Total Hrs",
        "Avg_Daily_Hrs": "Avg Daily Hrs",
        "Peak_Daily_Hrs": "Peak Daily Hrs",
        "Denial_Rate": "Denial %",
        "First_Seen": "First Seen",
        "Last_Seen": "Last Seen",
    })

    cols = ["User", "Hostname", "Software", "Active Days", "Events",
            "Checkouts", "Denials", "Denial %",
            "Total Hrs", "Avg Daily Hrs", "Peak Daily Hrs",
            "First Seen", "Last Seen"]
    return user_sw[cols].sort_values("Total Hrs", ascending=False)


def _build_top5_per_software(dash_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    """Top-5 users ranked by hours for each software, plus a consolidated Top-5.

    For software with no session-hour data (e.g. CATIA License, Ansys Peak),
    the ranking falls back to event count so every software still gets a
    meaningful Top-5 table.

    Returns ``{"Consolidated": df, "CATIA Usage": df, ...}``.
    """
    if dash_df.empty:
        return {}

    result: Dict[str, pd.DataFrame] = {}

    def _top5(sub: pd.DataFrame, label: str) -> pd.DataFrame:
        """Aggregate per user, rank, and return top 5."""
        agg = (sub.groupby("User")
               .agg(
                   Hostname=("Hostname", "first"),
                   Active_Days=("Date", "nunique"),
                   Events=("Events", "sum"),
                   Checkouts=("Checkouts", "sum"),
                   Denials=("Denials", "sum"),
                   Total_Hrs=("Daily Hrs", "sum"),
                   Avg_Daily_Hrs=("Daily Hrs", "mean"),
                   Peak_Daily_Hrs=("Daily Hrs", "max"),
               ).reset_index())

        agg["Total_Hrs"] = agg["Total_Hrs"].round(1)
        agg["Avg_Daily_Hrs"] = agg["Avg_Daily_Hrs"].round(2)
        agg["Peak_Daily_Hrs"] = agg["Peak_Daily_Hrs"].round(2)

        # Rank by hours first; if all zeros, rank by events
        has_hours = agg["Total_Hrs"].sum() > 0
        sort_col = "Total_Hrs" if has_hours else "Events"
        agg = agg.sort_values(sort_col, ascending=False).head(5).reset_index(drop=True)
        agg.insert(0, "Rank", range(1, len(agg) + 1))
        agg.insert(2, "Software", label)

        agg = agg.rename(columns={
            "Active_Days": "Active Days",
            "Total_Hrs": "Total Hrs",
            "Avg_Daily_Hrs": "Avg Daily Hrs",
            "Peak_Daily_Hrs": "Peak Daily Hrs",
        })
        return agg

    # Per-software Top 5
    for sw_name in sorted(dash_df["Software"].unique()):
        sw = dash_df[dash_df["Software"] == sw_name]
        if sw.empty:
            continue
        result[sw_name] = _top5(sw, sw_name)

    # Consolidated Top 5 (across all software)
    # Exclude "[Product]" entries (e.g. Ansys Peak products) from consolidated
    # ranking — they are not real users.
    real_users = dash_df[~dash_df["User"].str.startswith("[Product]", na=False)]
    if not real_users.empty:
        result["Consolidated"] = _top5(real_users, "All Software")
    else:
        result["Consolidated"] = _top5(dash_df, "All Software")

    return result


# ======================================================================
# Narrative builder  (plain English per software)
# ======================================================================

def _narrative(non_empty: Dict[str, pd.DataFrame]) -> str:
    """Build a multi-paragraph executive narrative."""
    parts = []

    if "ansys_peak" in non_empty:
        df = non_empty["ansys_peak"]
        summ = df[df.get("record_type", pd.Series(dtype=str)) == "summary"]
        n = int(summ["product"].nunique()) if "product" in summ.columns else 0
        if not summ.empty and "total_count" in summ.columns:
            work = summ.copy()
            work["total_count_num"] = pd.to_numeric(work["total_count"], errors="coerce").fillna(0)
            if "average_usage" in work.columns:
                work["average_usage_num"] = pd.to_numeric(work["average_usage"], errors="coerce").fillna(0)
            else:
                work["average_usage_num"] = 0

            top = work.nlargest(3, "total_count_num")
            names = ", ".join(
                f"{r['product']} ({float(r.get('average_usage_num', 0)):0.0%})" for _, r in top.iterrows()
            )
            parts.append(f"ANSYS: {n} products tracked. Top by usage: {names}.")
        else:
            parts.append(f"ANSYS: {n} products tracked.")

    if "catia_license" in non_empty:
        df = non_empty["catia_license"]
        denials = df[df["action"] == "LICENSE_DENIED"] if "action" in df.columns else pd.DataFrame()
        n = len(denials)
        if n:
            users = int(denials["user"].nunique()) if "user" in denials.columns else 0
            parts.append(f"CATIA: {n:,} license denials detected — {users} users impacted. Most denials due to all seats in use.")
        else:
            parts.append("CATIA: No license denials — all users served successfully.")

    if "cortona" in non_empty:
        df = non_empty["cortona"]
        outs = int((df["action"] == "OUT").sum()) if "action" in df.columns else 0
        dens = int((df["action"] == "DENIED").sum()) if "action" in df.columns else 0
        users = sorted(df[df["user"].notna()]["user"].unique()) if "user" in df.columns else []
        parts.append(f"CORTONA: {outs} checkouts, {dens} denials. Users: {', '.join(users)}.")

    if "matlab" in non_empty:
        df = non_empty["matlab"]
        errs = int((df["level"] == "E").sum()) if "level" in df.columns else 0
        warns = int((df["level"] == "W").sum()) if "level" in df.columns else 0
        parts.append(f"MATLAB: {'Healthy' if errs == 0 else f'{errs} errors found'}. {warns} warnings.")

    if "creo" in non_empty:
        parts.append(f"CREO: License entitlement data imported ({len(non_empty['creo'])} rows).")

    return " | ".join(parts) if parts else "No data to summarise."


# ======================================================================
# License Entitlements template builder
# ======================================================================

def _build_entitlements(non_empty: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    """Pre-fill a license entitlements template from detected products."""
    rows = []

    # Ansys products from peak data
    if "ansys_peak" in non_empty:
        df = non_empty["ansys_peak"]
        summ = df[df.get("record_type", pd.Series(dtype=str)) == "summary"]
        if not summ.empty and "product" in summ.columns:
            for _, r in summ.iterrows():
                rows.append({"Software": "Ansys", "Product / Feature": r["product"],
                             "License Count": "", "License Type": "",
                             "AMC End Date": "", "Contract Number": "", "Vendor": "ANSYS Inc."})

    # CATIA features from denials
    if "catia_license" in non_empty:
        feats = non_empty["catia_license"]["feature"].dropna().unique() if "feature" in non_empty["catia_license"].columns else []
        for f in sorted(set(feats))[:20]:
            rows.append({"Software": "CATIA", "Product / Feature": f,
                         "License Count": "", "License Type": "",
                         "AMC End Date": "", "Contract Number": "", "Vendor": "Dassault Systèmes"})

    # Cortona features
    if "cortona" in non_empty and "feature" in non_empty["cortona"].columns:
        for f in sorted(non_empty["cortona"]["feature"].dropna().unique()):
            rows.append({"Software": "Cortona 3D", "Product / Feature": f,
                         "License Count": "", "License Type": "",
                         "AMC End Date": "", "Contract Number": "", "Vendor": "Parallel Graphics"})

    if "matlab" in non_empty:
        rows.append({"Software": "MATLAB", "Product / Feature": "",
                     "License Count": "", "License Type": "",
                     "AMC End Date": "", "Contract Number": "", "Vendor": "MathWorks"})

    if "creo" in non_empty:
        rows.append({"Software": "Creo (PTC)", "Product / Feature": "",
                     "License Count": "", "License Type": "",
                     "AMC End Date": "", "Contract Number": "", "Vendor": "PTC Inc."})

    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ======================================================================
# Main report generator
# ======================================================================

def generate_report(data_by_type: Dict[str, pd.DataFrame], output_dir: Path) -> Path:
    """Generate a clean, formatted, insight-rich Excel report."""

    # Normalize input (drop Nones)
    data_by_type = {k: v for k, v in (data_by_type or {}).items() if v is not None}

    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"log_report_{ts}.xlsx"

    non_empty = {k: v for k, v in data_by_type.items() if not v.empty}
    if not non_empty:
        # Build a minimal workbook instead of crashing the GUI.
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard"
        row = _put_title(ws, 1, "Software License Log Report", "dashboard")
        row = _put_subtitle(ws, row, f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}")
        row = _put_subtitle(ws, row, "No usable rows were parsed from the selected log files.")
        row = _put_subtitle(ws, row, "Tip: confirm the selected file is a real log export (not an email .msg or report).")
        wb.save(output_path)
        return output_path

    from openpyxl import Workbook
    wb = Workbook()
    # Remove default sheet
    default_ws = wb.active
    if default_ws:
        default_ws.title = "_tmp_"

    # ==================================================================
    # 1. DASHBOARD
    # ==================================================================
    ws = wb.create_sheet("Dashboard", 0)
    row = _put_title(ws, 1, "Software License Log Report — Executive Dashboard", "dashboard")
    row = _put_subtitle(ws, row,
        f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  {_narrative(non_empty)}")

    row += 1
    dash = _build_dashboard(non_empty)
    if not dash.empty:
        row = _put_section(ws, row, "Overview — All Software", "dashboard", len(dash.columns))
        row = _put_table(ws, row, dash, "dashboard")

    # Record counts
    counts = pd.DataFrame([
        {"Log Type": k, "Records Parsed": f"{len(v):,}"} for k, v in non_empty.items()
    ])
    row = _put_section(ws, row, "Data Parsed", "dashboard", 2)
    row = _put_table(ws, row, counts, "dashboard")

    _auto_width(ws)
    ws.freeze_panes = "A4"

    # ==================================================================
    # 2. ANSYS
    # ==================================================================
    has_ansys = "ansys" in non_empty or "ansys_peak" in non_empty
    if has_ansys:
        ws = wb.create_sheet("Ansys")
        row = _put_title(ws, 1, "ANSYS — License Usage Analysis", "ansys")
        row = _put_subtitle(ws, row, _narrative({"ansys_peak": non_empty.get("ansys_peak", pd.DataFrame()),
                                                  "ansys": non_empty.get("ansys", pd.DataFrame())}))
        row += 1

        if "ansys_peak" in non_empty:
            for sec in _analyse_ansys_peak(non_empty["ansys_peak"]):
                ncols = len(sec["table"].columns) if not sec["table"].empty else 1
                row = _put_section(ws, row, sec["title"], "ansys", ncols)
                row = _put_table(ws, row, sec["table"], "ansys")

        if "ansys" in non_empty:
            for sec in _analyse_ansys_lm(non_empty["ansys"]):
                ncols = len(sec["table"].columns) if not sec["table"].empty else 1
                row = _put_section(ws, row, sec["title"], "ansys", ncols)
                row = _put_table(ws, row, sec["table"], "ansys")

        _auto_width(ws)
        ws.freeze_panes = "A3"

    # ==================================================================
    # 3. CATIA
    # ==================================================================
    has_catia = any(k.startswith("catia") for k in non_empty)
    if has_catia:
        ws = wb.create_sheet("CATIA")
        row = _put_title(ws, 1, "CATIA — License Usage Analysis", "catia")
        row = _put_subtitle(ws, row, _narrative({k: v for k, v in non_empty.items() if k.startswith("catia")}))
        row += 1

        if "catia_license" in non_empty:
            for sec in _analyse_catia(non_empty["catia_license"]):
                ncols = len(sec["table"].columns) if not sec["table"].empty else 1
                row = _put_section(ws, row, sec["title"], "catia", ncols)
                row = _put_table(ws, row, sec["table"], "catia")

        if "catia_token" in non_empty:
            for sec in _analyse_catia_token(non_empty["catia_token"]):
                ncols = len(sec["table"].columns) if not sec["table"].empty else 1
                row = _put_section(ws, row, sec["title"], "catia", ncols)
                row = _put_table(ws, row, sec["table"], "catia")

        if "catia_usage_stats" in non_empty:
            for sec in _analyse_catia_stats(non_empty["catia_usage_stats"]):
                ncols = len(sec["table"].columns) if not sec["table"].empty else 1
                row = _put_section(ws, row, sec["title"], "catia", ncols)
                row = _put_table(ws, row, sec["table"], "catia")

        _auto_width(ws)
        ws.freeze_panes = "A3"

    # ==================================================================
    # 4. CORTONA
    # ==================================================================
    has_cortona = "cortona" in non_empty or "cortona_admin" in non_empty
    if has_cortona:
        ws = wb.create_sheet("Cortona")
        row = _put_title(ws, 1, "CORTONA 3D — License Usage Analysis", "cortona")
        row = _put_subtitle(ws, row, _narrative({k: v for k, v in non_empty.items() if "cortona" in k}))
        row += 1

        if "cortona" in non_empty:
            for sec in _analyse_cortona(non_empty["cortona"]):
                ncols = len(sec["table"].columns) if not sec["table"].empty else 1
                row = _put_section(ws, row, sec["title"], "cortona", ncols)
                row = _put_table(ws, row, sec["table"], "cortona")

        if "cortona_admin" in non_empty:
            for sec in _analyse_cortona_admin(non_empty["cortona_admin"]):
                ncols = len(sec["table"].columns) if not sec["table"].empty else 1
                row = _put_section(ws, row, sec["title"], "cortona", ncols)
                row = _put_table(ws, row, sec["table"], "cortona")

        _auto_width(ws)
        ws.freeze_panes = "A3"

    # ==================================================================
    # 5. MATLAB
    # ==================================================================
    if "matlab" in non_empty:
        ws = wb.create_sheet("MATLAB")
        row = _put_title(ws, 1, "MATLAB — Service Health Analysis", "matlab")
        row = _put_subtitle(ws, row, _narrative({"matlab": non_empty["matlab"]}))
        row += 1

        for sec in _analyse_matlab(non_empty["matlab"]):
            ncols = len(sec["table"].columns) if not sec["table"].empty else 1
            row = _put_section(ws, row, sec["title"], "matlab", ncols)
            row = _put_table(ws, row, sec["table"], "matlab")

        _auto_width(ws)
        ws.freeze_panes = "A3"

    # ==================================================================
    # 6. CREO
    # ==================================================================
    if "creo" in non_empty:
        ws = wb.create_sheet("Creo")
        row = _put_title(ws, 1, "CREO (PTC) — License Information", "creo")
        row = _put_subtitle(ws, row, _narrative({"creo": non_empty["creo"]}))
        row += 1

        for sec in _analyse_creo(non_empty["creo"]):
            ncols = len(sec["table"].columns) if not sec["table"].empty else 1
            row = _put_section(ws, row, sec["title"], "creo", ncols)
            row = _put_table(ws, row, sec["table"], "creo")

        _auto_width(ws)
        ws.freeze_panes = "A3"

    # ==================================================================
    # 6B. NX SIEMENS
    # ==================================================================
    if "nx" in non_empty:
        ws = wb.create_sheet("NX")
        row = _put_title(ws, 1, "NX Siemens — License Usage Analysis", "dashboard")
        row = _put_subtitle(ws, row, "FlexLM/FlexNet OUT/IN/DENIED extracted from NX license debug logs.")
        row += 1

        nx_df = non_empty["nx"].copy()
        if "date" not in nx_df.columns:
            nx_df["date"] = nx_df.get("timestamp", "").astype(str).str[:10]

        # Overview
        outs = int((nx_df.get("action", pd.Series(dtype=str)) == "OUT").sum())
        ins = int((nx_df.get("action", pd.Series(dtype=str)) == "IN").sum())
        dens = int((nx_df.get("action", pd.Series(dtype=str)) == "DENIED").sum())
        users = int(nx_df["user"].nunique()) if "user" in nx_df.columns else 0
        feats = int(nx_df["feature"].nunique()) if "feature" in nx_df.columns else 0
        overview = pd.DataFrame([
            {"Metric": "Active Users", "Value": users},
            {"Metric": "Features", "Value": feats},
            {"Metric": "Total OUT", "Value": outs},
            {"Metric": "Total IN", "Value": ins},
            {"Metric": "Total DENIED", "Value": dens},
            {"Metric": "Denial Rate", "Value": f"{dens / (outs + dens) * 100:.1f}%" if (outs + dens) else "0%"},
        ])
        row = _put_section(ws, row, "📌 Overview", "dashboard", len(overview.columns))
        row = _put_table(ws, row, overview, "dashboard")

        dq = _data_quality_kpis(nx_df, ["user", "feature", "host"], ts_col="timestamp")
        row = _put_section(ws, row, "Data Quality & Coverage", "dashboard", len(dq.columns))
        row = _put_table(ws, row, dq, "dashboard")

        # Session reconstruction + concurrency sizing
        if all(c in nx_df.columns for c in ["timestamp", "action", "user", "feature"]):
            sess_src = nx_df[nx_df["action"].isin(["OUT", "IN"])].copy()
            group_cols = ["user", "feature"] + (["host"] if "host" in nx_df.columns else [])
            sessions = _build_sessions_from_out_in(sess_src, group_cols=group_cols)
            if not sessions.empty:
                p95 = _percentile(sessions["duration_min"], 0.95)
                p99 = _percentile(sessions["duration_min"], 0.99)
                sizing = pd.DataFrame([
                    {"Metric": "Reconstructed Sessions", "Value": f"{len(sessions):,}"},
                    {"Metric": "Avg Session (min)", "Value": round(float(sessions["duration_min"].mean()), 1)},
                    {"Metric": "P95 Session (min)", "Value": round(p95, 1) if p95 is not None else "N/A"},
                    {"Metric": "P99 Session (min)", "Value": round(p99, 1) if p99 is not None else "N/A"},
                    {"Metric": "Max Session (min)", "Value": round(float(sessions["duration_min"].max()), 1)},
                ])
                row = _put_section(ws, row, "Session Durations (reconstructed from OUT/IN)", "dashboard", len(sizing.columns))
                row = _put_table(ws, row, sizing, "dashboard")

                conc = _hourly_concurrency_from_sessions(sessions)
                if not conc.empty:
                    conc = conc.rename(columns={
                        "Hour": "Hour (0-23)",
                        "Avg_Concurrent": "Avg Concurrent",
                        "Peak_Concurrent": "Peak Concurrent",
                    })
                    note = pd.DataFrame([
                        {
                            "Item": "Concurrency is estimated",
                            "Details": "Computed from reconstructed OUT→IN sessions. If the log is missing IN events (or has resets), peak/avg concurrency may be understated.",
                        }
                    ])
                    row = _put_section(ws, row, "Note", "dashboard", len(note.columns))
                    row = _put_table(ws, row, note, "dashboard")

                    row = _put_section(ws, row, "Hourly Concurrency (estimated from sessions)", "dashboard", len(conc.columns))
                    row = _put_table(ws, row, conc, "dashboard")

        # Top Features
        if "feature" in nx_df.columns and "action" in nx_df.columns:
            feat = nx_df[nx_df["feature"].notna()].groupby("feature").agg(
                OUT=("action", lambda s: int((s == "OUT").sum())),
                IN=("action", lambda s: int((s == "IN").sum())),
                DENIED=("action", lambda s: int((s == "DENIED").sum())),
                Users=("user", pd.Series.nunique),
                Days=("date", pd.Series.nunique),
            ).reset_index()
            feat["Denial Rate"] = feat.apply(
                lambda r: f"{r['DENIED'] / (r['OUT'] + r['DENIED']) * 100:.1f}%" if (r["OUT"] + r["DENIED"]) > 0 else "0%",
                axis=1,
            )
            feat = feat.sort_values(["OUT", "DENIED"], ascending=[False, False]).head(15)
            feat.columns = ["Feature", "OUT", "IN", "DENIED", "Users", "Active Days", "Denial Rate"]

            row = _put_section(ws, row, "🔑 Top Features (by OUT)", "dashboard", len(feat.columns))
            row = _put_table(ws, row, feat, "dashboard")

        # Top Users
        if "user" in nx_df.columns and "action" in nx_df.columns:
            topu = nx_df[nx_df["user"].notna()].groupby("user").agg(
                OUT=("action", lambda s: int((s == "OUT").sum())),
                IN=("action", lambda s: int((s == "IN").sum())),
                DENIED=("action", lambda s: int((s == "DENIED").sum())),
                Features=("feature", pd.Series.nunique),
                Days=("date", pd.Series.nunique),
            ).reset_index()
            topu["Total"] = topu["OUT"] + topu["DENIED"]
            topu = topu.sort_values("Total", ascending=False).head(15)
            topu.columns = ["User", "OUT", "IN", "DENIED", "Features", "Active Days", "Total"]

            row = _put_section(ws, row, "👤 Top Users (by activity)", "dashboard", len(topu.columns))
            row = _put_table(ws, row, topu, "dashboard")

        # Rolling averages (daily)
        if "date" in nx_df.columns and "action" in nx_df.columns:
            daily = nx_df.groupby("date").agg(
                OUT=("action", lambda s: int((s == "OUT").sum())),
                IN=("action", lambda s: int((s == "IN").sum())),
                DENIED=("action", lambda s: int((s == "DENIED").sum())),
                Users=("user", pd.Series.nunique),
            ).reset_index()
            # Standardize column name
            if "date" in daily.columns:
                daily.rename(columns={"date": "Date"}, inplace=True)

            # Rolling window averages over 3/6/12 months using _window_avgs
            # NOTE: _window_avgs expects parallel date/value series.
            avgs = []
            for metric in ["OUT", "DENIED", "Users"]:
                vals = daily[metric]
                avg_map = _window_avgs(daily["Date"], vals)
                avgs.append({"Metric": f"Avg Daily {metric}", **{k: round(v, 2) for k, v in avg_map.items()}})
            avg_tbl = pd.DataFrame(avgs)

            row = _put_section(ws, row, "📊 Rolling Averages (Daily)", "dashboard", len(avg_tbl.columns))
            row = _put_table(ws, row, avg_tbl, "dashboard")

            # Monthly breakdown
            daily["date_dt"] = pd.to_datetime(daily["Date"], errors="coerce")
            daily["Month"] = daily["date_dt"].dt.to_period("M").astype(str)
            monthly = daily[daily["Month"].notna()].groupby("Month").agg(
                OUT=("OUT", "sum"),
                DENIED=("DENIED", "sum"),
                Users=("Users", "max"),
                Days=("OUT", "size"),
            ).reset_index()
            monthly["Avg OUT/Day"] = (monthly["OUT"] / monthly["Days"]).round(2)
            monthly["Avg DENIED/Day"] = (monthly["DENIED"] / monthly["Days"]).round(2)

            row = _put_section(ws, row, "📅 Monthly Activity", "dashboard", len(monthly.columns))
            row = _put_table(ws, row, monthly, "dashboard")

        # Hour-wise activity (0–23) — requires parseable timestamps
        if "timestamp" in nx_df.columns and "action" in nx_df.columns:
            tmp = nx_df.copy()
            tmp["_ts"] = pd.to_datetime(tmp["timestamp"], format="mixed", errors="coerce")
            tmp = tmp[tmp["_ts"].notna()].copy()
            if not tmp.empty:
                tmp["Hour"] = tmp["_ts"].dt.hour
                hour_tbl = tmp.groupby("Hour").agg(
                    OUT=("action", lambda s: int((s == "OUT").sum())),
                    IN=("action", lambda s: int((s == "IN").sum())),
                    DENIED=("action", lambda s: int((s == "DENIED").sum())),
                    Users=("user", pd.Series.nunique),
                ).reset_index()

                full = pd.DataFrame({"Hour": list(range(24))})
                hour_tbl = full.merge(hour_tbl, on="Hour", how="left").fillna(0)
                for c in ["OUT", "IN", "DENIED", "Users"]:
                    hour_tbl[c] = pd.to_numeric(hour_tbl[c], errors="coerce").fillna(0).astype(int)

                row = _put_section(ws, row, "Hourly Activity (0–23)", "dashboard", len(hour_tbl.columns))
                row = _put_table(ws, row, hour_tbl, "dashboard")

                # Hour-wise averages (per day)
                tmp["Date"] = tmp["_ts"].dt.strftime("%Y-%m-%d")
                per_day = tmp.groupby(["Date", "Hour"]).agg(
                    OUT=("action", lambda s: int((s == "OUT").sum())),
                    IN=("action", lambda s: int((s == "IN").sum())),
                    DENIED=("action", lambda s: int((s == "DENIED").sum())),
                    Users=("user", pd.Series.nunique),
                ).reset_index()
                avg_hour = _hourly_averages_from_daily(per_day, date_col="Date", hour_col="Hour")
                if not avg_hour.empty:
                    row = _put_section(ws, row, "Hourly Averages (per day)", "dashboard", len(avg_hour.columns))
                    row = _put_table(ws, row, avg_hour, "dashboard")

        # Hour-wise denials by feature (Top 10 features)
        if "timestamp" in nx_df.columns and "action" in nx_df.columns and "feature" in nx_df.columns:
            dens = nx_df[nx_df["action"] == "DENIED"].copy()
            dens["_ts"] = pd.to_datetime(dens["timestamp"], format="mixed", errors="coerce")
            dens = dens[dens["_ts"].notna()].copy()
            if not dens.empty:
                dens["Hour"] = dens["_ts"].dt.hour
                by_feat = (dens.groupby(["feature", "Hour"]).size()
                           .reset_index(name="DENIED"))
                top_feats = (by_feat.groupby("feature")["DENIED"].sum()
                             .sort_values(ascending=False).head(10).index)
                by_feat = by_feat[by_feat["feature"].isin(top_feats)].copy()
                by_feat.rename(columns={"feature": "Feature"}, inplace=True)
                by_feat = by_feat.sort_values(["Feature", "Hour"], ascending=[True, True])

                row = _put_section(ws, row, "Hourly Denials by Feature (Top 10)", "dashboard", len(by_feat.columns))
                row = _put_table(ws, row, by_feat, "dashboard")

        _auto_width(ws)
        ws.freeze_panes = "A3"

    # ==================================================================
    # 7. LICENSE ENTITLEMENTS TEMPLATE
    # ==================================================================
    ws = wb.create_sheet("License Entitlements")
    row = _put_title(ws, 1, "Current License Entitlements — Template", "template")
    row = _put_subtitle(ws, row,
        "ℹ️  Fill in your contract details below. This data comes from vendor agreements, not from logs.")
    row += 1

    ent = _build_entitlements(non_empty)
    if not ent.empty:
        ncols = len(ent.columns)
        row = _put_section(ws, row, "License Entitlements (Fill In Your Data)", "template", ncols)
        row = _put_table(ws, row, ent, "template")

    _auto_width(ws)
    ws.freeze_panes = "A3"
    
    # ==================================================================
    # 8. UNIFIED USER DASHBOARD
    # ==================================================================
    ws = wb.create_sheet("User Dashboard")
    row = _put_title(ws, 1, "Unified User Analytics Dashboard", "dashboard")
    row = _put_subtitle(ws, row, "Cross-platform user activity with session hours, rolling averages, checkouts & denials — all software types.")
    row += 1

    user_dash = _build_user_dashboard(non_empty)
    if not user_dash.empty:
        # ── A. Consolidated Software Summary ──
        sw_summary = _build_software_summary(user_dash)
        if not sw_summary.empty:
            row = _put_section(ws, row, "Software Summary (Consolidated)", "dashboard", len(sw_summary.columns))
            row = _put_table(ws, row, sw_summary, "dashboard")
            row += 1

        # ── B. Per-User Hour Summary (Consolidated) ──
        user_hr = _build_user_hour_summary(user_dash)
        if not user_hr.empty:
            row = _put_section(ws, row, f"User Summary — {len(user_hr)} users across all software", "dashboard", len(user_hr.columns))
            row = _put_table(ws, row, user_hr, "dashboard")
            row += 1

        # ── B2. Top 5 Users per Software + Consolidated ──
        top5_tables = _build_top5_per_software(user_dash)
        if top5_tables:
            # Consolidated Top 5 first
            if "Consolidated" in top5_tables and not top5_tables["Consolidated"].empty:
                t5 = top5_tables["Consolidated"]
                row = _put_section(ws, row, "🏆 Top 5 Users — All Software (by Hours)", "dashboard", len(t5.columns))
                row = _put_table(ws, row, t5, "dashboard")
                row += 1

            # Per-software Top 5
            for sw_name in sorted(k for k in top5_tables if k != "Consolidated"):
                t5 = top5_tables[sw_name]
                if t5.empty:
                    continue
                ranked_by = "Hours" if t5["Total Hrs"].sum() > 0 else "Events"
                row = _put_section(ws, row,
                    f"🏆 Top 5 Users — {sw_name} (by {ranked_by})", "dashboard", len(t5.columns))
                row = _put_table(ws, row, t5, "dashboard")
                row += 1

        # ── C. Per-Software Detail Sections ──
        for sw_name in sorted(user_dash["Software"].unique()):
            sw_data = user_dash[user_dash["Software"] == sw_name].copy()
            if sw_data.empty:
                continue
            ncols = len(sw_data.columns)
            n_users = sw_data["User"].nunique()
            n_rows = len(sw_data)
            total_hrs = sw_data["Daily Hrs"].sum()
            denials = int(sw_data["Denials"].sum())
            sub = f"{n_users} users · {n_rows} rows · {total_hrs:,.1f} hrs"
            if denials:
                sub += f" · {denials:,} denials"
            row = _put_section(ws, row, f"{sw_name}  —  {sub}", "dashboard", ncols)
            row = _put_table(ws, row, sw_data.reset_index(drop=True), "dashboard")
            row += 1
    else:
        row = _put_subtitle(ws, row, "No user-level activity data available to generate this dashboard.")

    _auto_width(ws)
    ws.freeze_panes = "A3"

    # ==================================================================
    # 8B. USERS (RAW)  -- simple, external-ready table starting at row 1
    # ==================================================================
    ws = wb.create_sheet("Users (Raw)")
    row = _put_title(ws, 1, "Users (Raw) — Unified per-user-per-day data", "dashboard", 13)
    row = _put_subtitle(ws, row,
        "Clean table export (starts immediately below). Use this sheet for pivot tables / external sharing.")
    row += 1

    if not user_dash.empty:
        row = _put_table_simple(ws, row, user_dash.reset_index(drop=True), "dashboard")
    else:
        row = _put_subtitle(ws, row, "No user-level activity data available.")

    _auto_width(ws)
    ws.freeze_panes = "A4"

    # ==================================================================
    # 9. CRITICAL USAGE SUMMARY (concise one-pager)
    # ==================================================================
    try:
        from reporting.critical_summary import build_critical_summary
        summaries = build_critical_summary(non_empty)
        if summaries:
            ws = wb.create_sheet("Critical Summary", 0)  # Insert as FIRST sheet
            row = _put_title(ws, 1, "Critical License Usage Summary", "dashboard", 8)
            row = _put_subtitle(ws, row,
                f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}  |  "
                f"Concise view of only the most critical metrics per software")
            row += 1

            for s in summaries:
                sw_name = s.get("software", "Unknown")
                vendor = s.get("vendor", "")
                date_range = s.get("date_range", "N/A")

                # Software header
                ncols = 6
                row = _put_section(ws, row,
                    f"{sw_name}  —  {vendor}  |  {date_range}", "dashboard", ncols)

                # Alerts
                alerts = s.get("alerts", [])
                if alerts:
                    for alert in alerts:
                        cell = ws.cell(row=row, column=1, value=f"  {alert}")
                        cell.font = Font("Calibri", 10, bold=True,
                            color="CC0000" if "🔴" in alert else
                                  ("CC8800" if "🟡" in alert else "228B22"))
                        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
                        row += 1
                    row += 1

                # Key Metrics as compact 2-column table
                metrics = s.get("key_metrics", {})
                if metrics:
                    metrics_df = pd.DataFrame([
                        {"Metric": k, "Value": str(v)} for k, v in metrics.items()
                    ])
                    row = _put_table(ws, row, metrics_df, "dashboard")

                # Top Users
                top_users = s.get("top_users", pd.DataFrame())
                if isinstance(top_users, pd.DataFrame) and not top_users.empty:
                    row = _put_section(ws, row, f"Top {len(top_users)} Users", "dashboard", len(top_users.columns))
                    row = _put_table(ws, row, top_users.reset_index(drop=True), "dashboard")

                # Top Features
                top_feat = s.get("top_features", pd.DataFrame())
                if isinstance(top_feat, pd.DataFrame) and not top_feat.empty:
                    row = _put_section(ws, row, f"Top {len(top_feat)} Features/Products", "dashboard", len(top_feat.columns))
                    row = _put_table(ws, row, top_feat.reset_index(drop=True), "dashboard")

                # Monthly Trend
                trend = s.get("monthly_trend", pd.DataFrame())
                if isinstance(trend, pd.DataFrame) and not trend.empty:
                    row = _put_section(ws, row, "Monthly Trend", "dashboard", len(trend.columns))
                    row = _put_table(ws, row, trend.reset_index(drop=True), "dashboard")

                row += 1  # gap between software sections

            _auto_width(ws)
            ws.freeze_panes = "A3"
    except Exception:
        pass  # Don't let critical summary failure break the main report

    # ==================================================================
    # Cleanup: remove temp default sheet
    # ==================================================================
    if "_tmp_" in wb.sheetnames:
        del wb["_tmp_"]

    wb.save(output_path)
    return output_path
